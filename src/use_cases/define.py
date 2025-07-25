import asyncio
import datetime
import os
import pickle
import time
from concurrent.futures import ThreadPoolExecutor

import openpyxl
import pandas as pd
import streamlit as st
from langchain.docstore.document import Document
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_openai import AzureChatOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

gpt4o = AzureChatOpenAI(
    azure_deployment="gpt-4o",
    openai_api_version="2024-02-15-preview",
    api_key="901cce4a4b4045b39727bf1ce0e36219",
    azure_endpoint="https://aifordata-azureopenai.openai.azure.com/",
    temperature=0,
    max_tokens=4096,
)

hf_embeddings = HuggingFaceEmbeddings(
    model_name="all-mpnet-base-v2",
    model_kwargs={"device": "cpu", "trust_remote_code": True},
)


# Initialise paths
DATA_DICTIONARY_PATH = "./input/data_dictionary_filtered.xlsx"
EMBEDDINGS_PATH = "./input/data_dictionary.pkl"


def read_schema_excel(schema_path):
    schema_df = pd.read_excel(schema_path, engine="openpyxl")

    schema_info = ""
    for schema in schema_df["Schema Name"].unique():
        schema_info += f"\nSCHEMA: {schema}"
        schema_tables = schema_df[schema_df["Schema Name"] == schema][
            "Table Name"
        ].unique()
        if "Column Name" in schema_df.columns:
            for table in schema_tables:
                columns = schema_df[
                    (schema_df["Schema Name"] == schema)
                    & (schema_df["Table Name"] == table)
                ]["Column Name"].tolist()
                # schema_info += f"\nTable: {table}\nColumns:\n{', '.join(column if (i+1) % 8 else column + '\n' for i, column in enumerate(columns))}\n"
                schema_info += (
                    f"\nTable: {table}\nColumns:\n"
                    + ",\n".join(
                        ", ".join(columns[i : i + 12])
                        for i in range(0, len(columns), 12)
                    )
                    + "\n"
                )
        else:
            for table in schema_tables:
                schema_info += f"\nTable: {table}"

    return schema_df, schema_info


def read_schema_in_chunks(schema_path):
    schema_df = pd.read_excel(schema_path, engine="openpyxl")

    schema_info_chunks = []
    current_chunk = ""
    chunk_size = 10
    for schema in schema_df["Schema Name"].unique():
        current_chunk = f"\nSCHEMA: {schema}\n"
        schema_tables = schema_df[schema_df["Schema Name"] == schema][
            "Table Name"
        ].unique()
        for i, table in enumerate(schema_tables):
            if i % chunk_size == 0 and i != 0:
                schema_info_chunks.append(current_chunk)
                current_chunk = ""
            current_chunk += f"\nTable: {table}\nColumns: "
            columns = schema_df[
                (schema_df["Schema Name"] == schema)
                & (schema_df["Table Name"] == table)
            ]["Column Name"].tolist()
            current_chunk += (
                ",\n".join(
                    ", ".join(columns[i : i + 12]) for i in range(0, len(columns), 12)
                )
                + "\n"
            )

        if current_chunk:
            schema_info_chunks.append(current_chunk)
            current_chunk = ""

    return schema_df, schema_info_chunks


def read_schema_in_chunks_and_batch(schema_path, chunk_size=1, batch_size=1):
    schema_df = pd.read_excel(schema_path, engine="openpyxl")

    schema_info_batches = []
    current_batch = []
    current_chunk = ""
    chunk_counter = 0

    for schema in schema_df["Schema Name"].unique():
        current_chunk = f"\nSCHEMA: {schema}\n"
        schema_tables = schema_df[schema_df["Schema Name"] == schema][
            "Table Name"
        ].unique()
        for table in schema_tables:
            if chunk_counter > chunk_size:
                current_batch.append(current_chunk)
                current_chunk = ""
                chunk_counter = 0
                if len(current_batch) == batch_size:
                    schema_info_batches.append(current_batch)
                    current_batch = []

            columns = (
                schema_df[
                    (schema_df["Schema Name"] == schema)
                    & (schema_df["Table Name"] == table)
                ]["Column Name"]
                .astype(str)
                .tolist()
            )
            for i in range(0, len(columns), 36):
                current_chunk += f"\nTable: {table}\nColumns: "
                current_chunk += (
                    ",\n".join(
                        ", ".join(columns[j : j + 12]) for j in range(i, i + 36, 12)
                    )
                    + "\n"
                )
                chunk_counter += 1

        if current_chunk:
            current_batch.append(current_chunk)
            current_chunk = ""
            if len(current_batch) == batch_size:
                schema_info_batches.append(current_batch)
                current_batch = []

    if current_batch:
        schema_info_batches.append(current_batch)

    return schema_df, schema_info_batches


def parallel_description(schema_description, schema_info, data_dictionary=[]):
    glossaries = []

    description_prompt = PromptTemplate.from_template(
        """
    You are a business analyst and a subject matter expert in the life insurance domain. I will provide the business purpose and description of a schema
    alongwith schema info containing a list of columns and the tables they belong to. The schema description may also contain additional information about the naming
    conventions and other nomenclatures followed in the schema. You have to precisely describe each and every column so that a business glossary can be created. 
    You will use your knowledge, augment it with your understanding of the information from the data dictionary, infer context and consider any additional information 
    present in the schema description for describing the tables and columns by using the schema description and info as provided below:

    The data dictionary consists of data under the following fields in the given order, seperated by '->':
    Data Domain -> Sub Domain -> Field Name -> Field Description
    The Field Description not only contains description of the field but also other relevant information encapsulating business specifications.

    Data Dictionary: {data_dictionary}
    Schema Description: {schema_description}\n
    Schema Info: {schema_info}\n

    The response should be in the following format:
    Table: table name\n
    Description: description of the table\n
    - Column 1 name: column description\n
    - Column 2 name: column description\n
    ...

    Do not include introductions, explanations and conclusions. Ensure the column descriptions are accurate.
    """
    )

    requests = [
        {
            "schema_description": schema_description,
            "schema_info": schema_info[i],
            "data_dictionary": (
                data_dictionary[i] if len(data_dictionary) > 0 else "N/A"
            ),
        }
        for i in range(len(schema_info))
    ]
    # print(requests)

    runnable = RunnablePassthrough()
    output_parser = StrOutputParser()
    describe = runnable | description_prompt | gpt4o | output_parser

    glossaries = asyncio.run(describe.abatch(requests))

    if not isinstance(glossaries, list):
        raise TypeError("Expected glossaries to be a list of strings")

    glossary = "  \n".join(glossaries)
    return glossary


def create_business_glossary_excel(schema_name, glossary):

    business_glossary_file_path = (
        schema_name.replace(".xlsx", "") + "_with_Business_Glossary.xlsx"
    )
    sheet_name = "Business Glossary"
    lines = glossary.strip().split("\n")
    tables = []
    current_table = {}
    for line in lines:
        if line.startswith("Table:"):
            if current_table:
                tables.append(current_table)
            current_table = {
                "table_name": line.split("Table:")[1].strip(),
                "description": "",
                "columns": [],
            }
        elif line.startswith("Description:"):
            current_table["description"] = line.split("Description:")[1].strip()
        elif line.startswith("-"):
            column_detail = line.split(":", 1)
            if len(column_detail) > 1:
                column_name = column_detail[0].strip("-").strip()
                column_description = column_detail[1].strip()
                current_table["columns"].append((column_name, column_description))
            else:
                print(f"Skipping improperly formatted line: {line.strip()}")

    if current_table:
        tables.append(current_table)

    rows = []
    for table in tables:
        table_name = table["table_name"]
        table_description = table["description"]
        columns = table["columns"]
        for column_name, column_description in columns:
            rows.append(
                [table_name, table_description, column_name, column_description]
            )

    schema_df = pd.DataFrame(
        rows,
        columns=[
            "Table Name",
            "Table Description",
            "Column Name",
            "Column Description",
        ],
    )
    schema_df.to_excel(
        f"./output/business glossary/{business_glossary_file_path}",
        sheet_name=sheet_name,
        index=False,
    )
    # ----------------------------------

    workbook = load_workbook(
        f"./output/business glossary/{business_glossary_file_path}"
    )
    sheet = workbook[sheet_name]

    start_row = 2  # Assuming the first row is the header
    col_a_letter = "A"
    col_b_letter = "B"

    row = start_row
    while row <= sheet.max_row:
        cell_a = sheet[f"{col_a_letter}{row}"]
        cell_b = sheet[f"{col_b_letter}{row}"]

        if cell_a.value is not None:
            merge_start = row
            merge_end = row

            while (
                merge_end < sheet.max_row
                and sheet[f"{col_a_letter}{merge_end + 1}"].value == cell_a.value
            ):
                merge_end += 1

            if merge_end > merge_start:
                sheet.merge_cells(
                    start_row=merge_start,
                    start_column=cell_b.column,
                    end_row=merge_end,
                    end_column=cell_b.column,
                )
                merged_cell = sheet.cell(row=merge_start, column=cell_b.column)
                merged_cell.alignment = Alignment(wrapText=True, vertical="top")

            row = merge_end + 1
        else:
            row += 1

    header_fill = PatternFill(
        start_color="D0E8FA", end_color="D0E8FA", fill_type="solid"
    )
    for cell in sheet[1]:
        cell.fill = header_fill

    workbook.save(f"./output/business glossary/{business_glossary_file_path}")


# Streamlit Interface
if "schema_info" not in st.session_state:
    st.session_state["schema_info"] = ""

if "schema_info_batches" not in st.session_state:
    st.session_state["schema_info_batches"] = []

if "schema_df" not in st.session_state:
    st.session_state["schema_df"] = None

if "dd" not in st.session_state:
    st.session_state["dd"] = []

if "glossary" not in st.session_state:
    st.session_state["glossary"] = ""

if "glossary_continuation" not in st.session_state:
    st.session_state["glossary_continuation"] = ""


st.set_page_config(
    layout="wide",
    page_title="Automatic Generation Business Glossary",
    page_icon="📖",
)

st.subheader(
    "Automatic Creation of Business Glossary using EY's Expertise and Large Language Models 📖"
)
st.markdown(
    """
The accelerator supports teams in cataloging source systems which may often contain ambiguous names for entities and attributes. 
It resolves such ambiguities by leveraging EY's extensive experience in the financial and life insurance sector. It accepts the 
path to any excel sheet containing a list of table names and column names and a brief description of the source system where you
may include further specifications regarding updates and naming conventions. The accelerator relies on a RAG pipeline to a knowledge 
base and a Large Language Model.

The outcome is an excel sheet containing table and column descriptions, which can be found in the **output/business glossary** directory.

----
"""
)

# Input field for Excel sheet path
schema_path = st.text_input("Enter the path to the schema Excel sheet:")
schema_path = schema_path.strip('"')
schema_name = os.path.basename(schema_path)

# Text area for schema description
schema_description = st.text_area("Enter the schema description:")

if schema_path and schema_description:

    schema_df, schema_info_batches = read_schema_in_chunks_and_batch(schema_path)
    st.session_state["schema_info_batches"] = schema_info_batches
    st.session_state["schema_df"] = schema_df

    if st.button("Describe") and "schema_info_batches" in st.session_state:
        for batch_index, batch in enumerate(st.session_state["schema_info_batches"]):
            try:
                with st.spinner(
                    f"Generating Business Glossary for {len(st.session_state['schema_info_batches'])} batches"
                ):
                    # st.text(batch)
                    glossary = parallel_description(
                        schema_description, batch, st.session_state["dd"]
                    )
                    st.session_state["glossary_continuation"] = f"  \n{glossary}"
                    with open(
                        f"./output/business glossary/{schema_name.replace('.xlsx', 'Business Glossary.txt')}",
                        "a",
                    ) as text_file:
                        text_file.write(st.session_state["glossary_continuation"])
                    st.session_state["glossary"] += st.session_state[
                        "glossary_continuation"
                    ]
                st.markdown(st.session_state["glossary_continuation"])
                time.sleep(5)
            except:
                continue

    # Button to get Business Glossary Excel sheet
    if (
        st.button("Get Business Glossary Excel Sheet")
        and "glossary" in st.session_state
    ):
        create_business_glossary_excel(schema_name, st.session_state["glossary"])
        st.success("Business Glossary Excel sheet created successfully!")
