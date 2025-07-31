import os
import re

import openpyxl
import pandas as pd
import streamlit as st
import tiktoken
from graphviz import Digraph, Source
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_openai import AzureChatOpenAI
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill

from src.config.settings import settings

llm = AzureChatOpenAI(
    azure_deployment="gpt-4o",
    openai_api_version="2024-02-15-preview",
    api_key="901cce4a4b4045b39727bf1ce0e36219",
    azure_endpoint="https://aifordata-azureopenai.openai.azure.com/",
    temperature=0,
    max_tokens=4096,
)


# Set path for executables
os.environ["PATH"] += os.pathsep + r"C:\Program Files\Graphviz\bin"


def filter_table(table):
    rows = table.strip().split("\n")
    expected_parts = 15
    table_new_list = rows[:2]

    filtered_rows = rows[2:]
    for row in filtered_rows[::-1]:
        if "------" in row:
            return table

        parts = row.split("|")
        if len(parts) < expected_parts or row.count("|") < expected_parts - 1:
            filtered_rows.remove(row)
        else:
            break

    table_new_list.extend(filtered_rows)
    table_new = "\n".join(table_new_list)
    return table_new


def process_table(table):
    rows = []
    try:
        for line in table.strip().split("\n")[2:]:
            if line.count("|") == 14 or line.count("|") == 15:
                columns = line.split("|")
                row = {
                    "Layer": columns[1].strip(),
                    "Description": columns[2].strip(),
                    "Table": columns[3].strip(),
                    "Table Description": columns[4].strip(),
                    "Column": columns[5].strip(),
                    "Data Type": columns[6].strip(),
                    "Column Description": columns[7].strip(),
                    "Join Table": columns[8].strip() if columns[8] != "" else "S/B",
                    "Join Column": columns[9].strip() if columns[9] != "" else "S/B",
                    "Join Statement": (
                        columns[10].strip() if columns[10] != "" else "S/B"
                    ),
                    "Filter Table": columns[11].strip() if columns[11] != "" else "S/B",
                    "Filter Column": (
                        columns[12].strip() if columns[12] != "" else "S/B"
                    ),
                    "Filter Statement": (
                        columns[13].strip() if columns[13] != "" else "S/B"
                    ),
                    "Derivation Purpose": (
                        columns[14].strip() if columns[14] != "" else "S/B"
                    ),
                }
                rows.append(row)

        df = pd.DataFrame(rows)
        df_filled = df.replace("S/B", pd.NA)
        df_filled = df_filled.ffill()
        df_filled["Join Table"] = df_filled["Table"].str.strip("@").str.upper()
        df_filled["Filter Table"] = df_filled["Table"].str.strip("@").str.upper()
        df_filled["Table"] = df_filled["Table"].str.strip("@").str.upper()

        return df, df_filled

    except Exception as e:
        st.error(e)


def get_glossary(plsql_script, file):
    extract_tables_prompt = PromptTemplate.from_template(
        """
    Extract all the source table names from the following stored procedure:
    {plsql}

    Respond with a comma separated list of the source table names only.
    """
    )
    runnable = RunnablePassthrough()
    output_parser = StrOutputParser()

    extract_tables_chain = runnable | extract_tables_prompt | llm | output_parser
    tables = extract_tables_chain.invoke({"plsql": plsql_script})
    tables = tables[tables.find(":") + 1 :].strip()

    glossary_df = pd.read_excel(file)
    glossary_list = []
    for table in tables.split(", "):
        plsql_table_glossary = glossary_df[
            any(re.split(r"[_\.]", table.upper()) in glossary_df["Table Name"])
        ][["Table Name", "Table Description", "Column Name", "Column Description"]]
        if not plsql_table_glossary.empty:
            glossary_list.append(plsql_table_glossary)

    result = ""
    if len(glossary_list) > 0:
        plsql_glossary = pd.concat(glossary_list, ignore_index=True)
        for table in plsql_glossary["Table Name"].unique():
            result += f"{table}\n{plsql_glossary[plsql_glossary['Table Name'] == table]['Table Description'].iloc[0]}\n"
            columns = plsql_glossary[plsql_glossary["Table Name"] == table]
            for _, row in columns.iterrows():
                result += f"{row['Column Name']}: {row['Column Description']}\n"
            result += "\n"

    # print(result)
    return result


def analyse_sp(plsql_script, context="N/A", dialect="T-SQL"):
    if dialect == "T-SQL":
        analysis_prompt = PromptTemplate.from_template(
            """
        You are a data engineer with expertise in understanding optimized, enterprise-grade PL/SQL procedures with a focus on the BFSI sector. Analyze the provided PL/SQL script 
        which is being used to transform data from source systems build reports in the life insurance domain. Create a thorough and comprehensive table that captures the necessary details for both 
        business analysts and data engineers. Optionally, you may be provided with a business glossary of the source system. Use the business glossary of the data source to augment your analysis.
        The table should include information about multiple layers of transformations, tables and columns involved in each section, data types, key constraints, description of each and every column, 
        join tables and columns, join conditions, filter conditions, transformation statements, and the purpose of applying the joins, filters and transformations. Follow these steps to sequentially 
        perform the analysis and populate the table:

        1. Analyse each and every section of the PL/SQL procedure. The PL/SQL procedure consists of multiple layers of data transformation, designed to optimize query performance. 
        Determine the purpose of each section in the data engineering and report building process. You will utilize this analysis in naming and describing the sections as well as 
        in documenting the tables involved in the section. Ensure that each section has only one concise and high-level description documented under "Section Description".

        2. Identify the variables, tables CTEs and columns in the PL/SQL procedure. Analyse the insert select, from and join statements, aggregations, select case when and other filter conditions.
        You will use this analysis to document the table name, the columns belonging to the table and their data types in a hierarchial manner of data flow. You will also use this analysis to determine the purpose
        of the table and column in the report building process.

        3. Since this is a multi-stage transformation process, multiple tables and columns are derived with join statements, complex mathematical formulae, case select conditions, 
        aggregations, filters, etc. The derived tables and columns may further be transformed into new tables and columns in a similar fashion. You will identify the statements pertaining 
        to deriving a table and its column. You will analyse these statements and segment them into the parts responsible for performing join operations and parts responsible for applying 
        business logic such as filters based on case when statements, aggregations, etc. If a statement does not contain any join operations, you will consider it to not be applicable.
        You will map aliases to their actual table names and identify all the columns and tables being used in both sides of the various join operation, subsequent boolean operators and subqueries.
        You will identify all the tables and columns being used in insert, from, aggregation, select when and other filter conditions for deriving a table and its columns. 
        You will analyse these statements and use the analyses from steps 1 and 2 to comprehend the derivation logic and the purpose of the derivation logic in the overall report building process. 
        You will use these analyses to accurately document every table and column being invoked in the join part of the statements under join table and join column seperated by commas and you will only 
        document the join operations in the statements under join statements. 
        You will document the tables and columns invoked in case when, aggregation and other filter conditions for deriving a particular table and column under derivation table, derivation column, separated 
        by commas, truncated versions of the derivation statements excluding the join statements, under derivation statement and the purpose and description of the derivation logic being applied to derive them.

        For example, in the statement "insert @lins select a.CHDrnum, case when b.chdrnum is null then 'O' else 'R' end, a.ptrneff,a.PTRNEFF from @PTRN1 a left join PTRNPF b on a.chdrnum = b.chdrnum and a.ptrneff = b.PTRNEFF and 
        b.validflag !='2' and b.batctrcde = 'B522'", the join tables are "@PTRN1, PTRNPF" which have been mapped to aliases 'a' and 'b' respectively, join columns are "@PTRN1.chdrnum, PTRNPF.chdrnum, @PTRN1.ptrneff, PTRNPF.PTRNEFF", 
        'PTRNPF.validflag' and 'PTRNPF.batctrcde' will not be included in the join columns since they filter conditions, join statement is "from @PTRN1 a left join PTRNPF b on a.chdrnum = b.chdrnum and a.ptrneff = b.PTRNEFF", 
        derivation tables are "@PTRN1, PTRNPF", derivation columns are "PTRNPF.chdrnum, PTRNPF.validflag, PTRNPF.batctrcde" and derivation statement is "insert @lins select a.CHDrnum, case when b.chdrnum is null then 'O' else 'R' end..."

        4. Identify cursor declarations (e.g., CURSOR cursor_name IS SELECT ...). If a cursor is used to populate a table or derive a column, document the cursor's name and query under 
        "Derivation Table" and "Derivation Statements" respectively along with an explanation of its purpose in "Purpose of Derivation.
        For each loop (e.g., FOR record IN cursor_name LOOP ... END LOOP;), document the loop variable and the cursor being iterated over, under "Column" and "Table" respectively.
        Analyze the set of operations within the loop that process the rows fetched from the cursor. Document these operations within the "Derivation Statements", focusing 
        on how they contribute to the derivation of the table or column.
        
        For example, if a cursor fetches policy data and a loop calculates the total premium for each policy, and this total is used in a column, the documentation for that column might include:
        Derivation Table: "policy_cursor"
        Derivation Column: "policy_record"
        Derivation Statements: "... CURSOR policy_cursor IS SELECT policy_id, premium_amount FROM policy_table WHERE status = 'Active'; FOR policy_record IN policy_cursor LOOP total_premium := total_premium + policy_record.premium_amount; END LOOP; ..."
        Purpose of Derivation: "... Calculate the total premium for each active policy using a cursor and loop..."

        5. Identify function declarations (e.g., function getfxdcrncyunits(fxdcrncycode varchar2, varcrncycode varchar2, rate_code varchar2, asondate date)). Document the functions under "Table" alongwith
        any resulting variables from the function's logic under "Column". You will also document the join queries, derivation logic and purpose for their derivation under the relevant columns.

        6. Since this may a PL/SQL procedure to provide data for a reporting dashboard, you will also analyse the section of the procedure which applies SELECT, UNION, JOIN and other such operations on the tables and columns derived 
        from the various stages of the transformation to procure the data for the report. These elements may not have table and column names, so you will assign a table and column name to them and an accurate description when documenting this analysis.
        Suitable table names may be report, dbms_output, etc

        7. Finally, you will populate the table thoroughly with the technical specifics and clear, contextual descriptions you will obtain from the above steps. The table should contain the following headers: 
        | Section Name | Section Description | Table Name | Table Description | Column Name | Data Type | Column Description | Join Table | Join Column | Join Statements | Derivation Table | Derivation Column | Derivation Statements | Purpose of Derivation |\n
        | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |\n
        You will write N/A where values are not applicable.
        Ensure that all the rows begin and end with |.
        Truncate join statements and derivation statements to 8 keywords then add it to the documentation with "..." to denote continuation.
        If a Section Name and consequently Section Description pertains to multiple tables and their columns, write S/B for subsequent instances. 
        If a Table Name and consequently, Table Description pertains to multiple columns, write S/B for repeating instances.
        If a Join Table, Join Column, Join Statement, Derivation Table, Derivation Column, Derivation Statement and consequently the Purpose of Derivation pertains to multiple Column Names, fill the row 
        for the first instance with the derivation statement as instructed above, and write "S/B" for subsequent instances.

        Following is the PL/SQL script:
        {plsql_script}

        Following is the business glossary of the source system:
        {context}

        The resulting table will provide a detailed overview of the PL/SQL script's operations, data flow, and business logic. It should serve as a reference guide for data engineers 
        to understand the exact queries and operations applied, and for business analysts to grasp the purpose and functionality of each section, element and statements within the script.
        Perform the documentation thoroughly and ensure no part of your analysis is excluded from the table as this documentation is highly critical. Perform as much of the analysis as you can. 
        Respond with as many complete rows of the table as you can create. if the table does not contain analysis and documentation of the entire PL/SQL script I will prompt you to continue 
        and then you will continue populating the table. Do not include introductions, labels, backticks and conclusions. Only respond with the table. Each row of the table should end with a newline character.
        If the analysis is complete for the entire stored procedure and the table is complete, do not start over from the beginning and end with ------. 
        """
        )

        continue_analysis_prompt = PromptTemplate.from_template(
            """
        You are a data engineer with expertise in understanding optimized, enterprise-grade PL/SQL procedures with a focus on the BFSI sector. You were tasked with analyzing the 
        provided PL/SQL script which is being used to build reports in the life insurance domain. You created part of a thorough and comprehensive table that captures the 
        necessary details for both business analysts and data engineers by analysing the PL/SQL script and augmenting your analysis with the business glossary of the data source. The table includes 
        information about multiple layers of transformations, tables and columns involved in each section, data types, key constraints, description of each and every, join tables and columns, join 
        conditions, filter conditions, transformation statements, and the purpose of applying the joins, filters and transformations. 

        Following is the table you have created so far:
        {table}

        Analyse the latest rows of the table created so far and determine where to analyse the PL/SQL procedure from. Follow these instructions to continue creating the table:
        1. Analyse each and every section of the PL/SQL procedure. The PL/SQL procedure consists of multiple layers of data transformation, designed to optimize query performance. 
        Determine the purpose of each section in the data engineering and report building process. You will utilize this analysis in naming and describing the sections as well as 
        in documenting the tables involved in the section. Ensure that each section has only one concise and high-level description documented under "Section Description".

        2. Identify the variables, tables CTEs and columns in the PL/SQL procedure. Analyse the insert select, from and join statements, aggregations, select case when and other filter conditions.
        You will use this analysis to document the table name, the columns belonging to the table and their data types in a hierarchial manner of data flow. You will also use this analysis to 
        determine the purpose of the table and column in the report building process.

        3. Since this is a multi-stage transformation process, multiple tables and columns are derived with join statements, complex mathematical formulae, case select conditions, 
        aggregations, filters, etc. The derived tables and columns may further be transformed into new tables and columns in a similar fashion. You will identify the statements pertaining 
        to deriving a table and its column. You will analyse these statements and segment them into the parts responsible for performing join operations and parts responsible for applying 
        business logic such as filters based on case when statements, aggregations, etc. If a statement does not contain any join operations, you will consider it to not be applicable.
        You will map aliases to their actual table names and identify all the columns and tables being used in both sides of the various join operation, subsequent boolean operators and subqueries.
        You will identify all the tables and columns being used in insert, from, aggregation, select when and other filter conditions for deriving a table and its columns. 
        You will analyse these statements and use the analyses from steps 1 and 2 to comprehend the derivation logic and the purpose of the derivation logic in the overall report building process. 
        You will use these analyses to accurately document every table and column being invoked in the join part of the statements under join table and join column seperated by commas and you will only 
        document the join operations in the statements under join statements. 
        You will document the tables and columns invoked in case when, aggregation and other filter conditions for deriving a particular table and column under derivation table, derivation column, separated 
        by commas, truncated versions of the derivation statements excluding the join statements, under derivation statement and the purpose and description of the derivation logic being applied to derive them.

        For example, in the statement "insert @lins select a.CHDrnum, case when b.chdrnum is null then 'O' else 'R' end, a.ptrneff,a.PTRNEFF from @PTRN1 a left join PTRNPF b on a.chdrnum = b.chdrnum and a.ptrneff = b.PTRNEFF and 
        b.validflag !='2' and b.batctrcde = 'B522'", the join tables are "@PTRN1, PTRNPF" which have been mapped to aliases 'a' and 'b' respectively, join columns are "@PTRN1.chdrnum, PTRNPF.chdrnum, @PTRN1.ptrneff, PTRNPF.PTRNEFF", 
        'PTRNPF.validflag' and 'PTRNPF.batctrcde' will not be included in the join columns since they filter conditions, join statement is "from @PTRN1 a left join PTRNPF b on a.chdrnum = b.chdrnum and a.ptrneff = b.PTRNEFF", 
        derivation tables are "@PTRN1, PTRNPF", derivation columns are "PTRNPF.chdrnum, PTRNPF.validflag, PTRNPF.batctrcde" and derivation statement is "insert @lins select a.CHDrnum, case when b.chdrnum is null then 'O' else 'R' end..."

        4. Cursor and Loop Handling:
        Identify cursor declarations (e.g., CURSOR cursor_name IS SELECT ...). If a cursor is used to populate a table or derive a column, document the cursor's name and query under "Table" and "Derivation Statements" respectively.
        along with an explanation of its purpose in "Purpose of Derivation.
        For each loop (e.g., FOR record IN cursor_name LOOP ... END LOOP;), document the loop variable and the cursor being iterated over, under "Column" and "Table" respectively.
        Analyze the set of operations within the loop that process the rows fetched from the cursor. Document these operations within the "Derivation Statements", focusing 
        on how they contribute to the derivation of the table or column.
        
        For example, if a cursor fetches policy data and a loop calculates the total premium for each policy, and this total is used in a column, the documentation for that column might include:
        Derivation Table: "policy_cursor"
        Derivation Column: "policy_record"
        Derivation Statements: "... CURSOR policy_cursor IS SELECT policy_id, premium_amount FROM policy_table WHERE status = 'Active'; FOR policy_record IN policy_cursor LOOP total_premium := total_premium + policy_record.premium_amount; END LOOP; ..."
        Purpose of Derivation: "... Calculate the total premium for each active policy using a cursor and loop..."

        5. Since this may a PL/SQL procedure to provide data for a reporting dashboard, you will also analyse the section of the procedure which applies SELECT, UNION, JOIN and other such operations on the tables and columns derived 
        from the various stages of the transformation to procure the data for the report. These elements may not have table and column names, so you will assign a table and column name to them and an accurate description when documenting this analysis.
        Suitable table names may be report, dbms_output, etc.

        6. Identify function declarations (e.g., function getfxdcrncyunits(fxdcrncycode varchar2, varcrncycode varchar2, rate_code varchar2, asondate date)). Document the functions under "Table" alongwith
        any resulting variables from the function's logic under "Column". You will also document the join queries, derivation logic and purpose for their derivation under the relevant columns.

        7. Finally, you will populate the table thoroughly with the technical specifics and clear, contextual descriptions you will obtain from the above steps. The table should contain the following headers: 
        | Section Name | Section Description | Table Name | Table Description | Column Name | Data Type | Column Description | Join Table | Join Column | Join Statements | Derivation Table | Derivation Column | Derivation Statements | Purpose of Derivation |\n
        | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |\n
        You will write N/A where values are not applicable.
        Ensure that all the rows begin and end with |.
        Truncate join statements and derivation statements to 8 keywords then add it to the documentation with "..." to denote continuation.
        If a Section Name and consequently Section Description pertains to multiple tables and their columns, write S/B for repeating instances. 
        If a Table Name and consequently, Table Description pertains to multiple columns, write S/B for repeating instances.
        If a Join Table, Join Column, Join Statement, Derivation Table, Derivation Column, Derivation Statement and consequently the Purpose of Derivation pertains to multiple Column Names, fill the row 
        for the first instance with the derivation statement as instructed above, and write "S/B" for subsequent instances.

        Following is the PL/SQL script:
        {plsql_script}

        Following is the business glossary of the data source:
        {context} 

        The resulting table will provide a detailed overview of the PL/SQL script's operations, data flow, and business logic. It should serve as a reference guide for data engineers 
        to understand the exact queries and operations applied, and for business analysts to grasp the purpose and functionality of each section, element and statements within the script.
        Perform a complete analysis. Perform the documentation thoroughly. Do not include introductions, labels, backticks and conclusions. Dont include headers. Only respond with the new 
        rows of the table. Each row should end with a newline character. If the table is complete or has nearly been completed, i.e only one to three rows are remaining, respond with the remaining 
        rows if any, do not start over and end your response with ------.
        """
        )

    if dialect == "Spark SQL":
        analysis_prompt = PromptTemplate.from_template(
            """        
        You are a data engineer with expertise in understanding optimized, enterprise-grade Spark SQL and PySpark transformations with a focus on the BFSI sector. Analyze the provided Spark SQL or PySpark script which is being used to
        transform data from source systems to build reports in the life insurance domain. Create a thorough and comprehensive table that captures the necessary details for both business analysts and data engineers. Optionally, you may 
        be provided with a business glossary of the source system. Use the business glossary of the data source to augment your analysis. The table should include information about multiple layers of transformations, tables and columns 
        involved in each section, data types, key constraints, description of each and every column, join tables and columns, join conditions, filter conditions, transformation statements, and the purpose of applying the joins, filters 
        and transformations. Follow these steps to sequentially perform the analysis and populate the table:

        1. Analyse each and every section of the Spark script (Spark SQL or PySpark). The script consists of multiple layers of data transformation, designed to optimize query performance. Determine the purpose of each section in the data 
        engineering and report building process. You will utilize this analysis in naming and describing the sections as well as in documenting the tables/DataFrames involved in the section. Ensure that each section has only one concise and 
        high-level description documented under "Section Description".

        2. Identify the variables, tables/DataFrames, and columns in the Spark script. Analyse the select, from, and join statements, aggregations, CASE WHEN clauses, and other filter conditions. You will use this analysis to document the 
        table/DataFrame name, the columns belonging to the table/DataFrame and their data types in a hierarchical manner of data flow. You will also use this analysis to determine the purpose of the table/DataFrame and column in the report 
        building process.

        3. Since this is a multi-stage transformation process, multiple tables/DataFrames and columns are derived with join statements, complex mathematical formulae, CASE WHEN conditions, aggregations, filters, etc. The derived tables/DataFrames 
        and columns may further be transformed into new tables/DataFrames and columns in a similar fashion. You will identify the statements pertaining to deriving a table/DataFrame and its column. You will analyse these statements and segment them 
        into the parts responsible for performing join operations and parts responsible for applying business logic such as filters based on CASE WHEN statements, aggregations, etc. If a statement does not contain any join operations, you will consider 
        it to not be applicable. You will map aliases to their actual table/DataFrame names and identify all the columns and tables/DataFrames being used in both sides of the various join operations, subsequent boolean operators and subqueries. You will 
        identify all the tables/DataFrames and columns being used in select, from, aggregation, CASE WHEN and other filter conditions for deriving a table/DataFrame and its columns. You will analyse these statements and use the analyses from steps 1 and 2 
        to comprehend the derivation logic and the purpose of the derivation logic in the overall report building process. You will use these analyses to accurately document every table/DataFrame and column being invoked in the join part of the statements 
        under join table and join column separated by commas and you will only document the join operations in the statements under join statements. You will document the tables/DataFrames and columns invoked in CASE WHEN, aggregation and other filter 
        conditions for deriving a particular table/DataFrame and column under derivation table, derivation column, separated by commas, truncated versions of the derivation statements excluding the join statements, under derivation statement and the purpose
        and description of the derivation logic being applied to derive them.
        For example, in the Spark SQL statement "INSERT INTO lins SELECT a.CHDrnum, CASE WHEN b.chdrnum IS NULL THEN 'O' ELSE 'R' END, a.ptrneff FROM PTRN1 a LEFT JOIN PTRNPF b ON a.chdrnum = b.chdrnum AND a.ptrneff = b.PTRNEFF WHERE b.validflag != '2' AND b.batctrcde = 'B522'", or its PySpark equivalent:
        lins = spark.sql("SELECT a.CHDrnum, CASE WHEN b.chdrnum IS NULL THEN 'O' ELSE 'R' END, a.ptrneff FROM PTRN1 a LEFT JOIN PTRNPF b ON a.chdrnum = b.chdrnum AND a.ptrneff = b.PTRNEFF WHERE b.validflag != '2' AND b.batctrcde = 'B522'")
        the join tables are "PTRN1, PTRNPF" which have been mapped to aliases 'a' and 'b' respectively, join columns are "PTRN1.chdrnum, PTRNPF.chdrnum, PTRN1.ptrneff, PTRNPF.PTRNEFF", 'PTRNPF.validflag' and 'PTRNPF.batctrcde' will not be included in the 
        join columns since they filter conditions, join statement is "FROM PTRN1 a LEFT JOIN PTRNPF b ON a.chdrnum = b.chdrnum AND a.ptrneff = b.PTRNEFF", derivation tables are "PTRN1, PTRNPF", derivation columns are "PTRNPF.chdrnum, PTRNPF.validflag, PTRNPF.batctrcde" 
        and derivation statement is "INSERT INTO lins SELECT a.CHDrnum, CASE WHEN b.chdrnum IS NULL THEN 'O' ELSE 'R' END...".

        4. If DataFrames are created using transformations like groupBy, agg, withColumn, document the DataFrame's name under "Derivation Table". Document the transformations used to create the DataFrame under "Derivation Statements," focusing on how they 
        contribute to the derivation of the table/DataFrame or column.
        For example, if a DataFrame policy_summary is created by grouping by policy_id and summing premium_amount:
        Spark SQL:
        CREATE OR REPLACE TEMP VIEW policy_summary AS
        SELECT policy_id, SUM(premium_amount) as total_premium FROM policy_table WHERE status = 'Active' GROUP BY policy_id

        PySpark:
        policy_summary = policy_table.filter("status = 'Active'").groupBy("policy_id").agg(sum("premium_amount").alias("total_premium"))
        policy_summary.createOrReplaceTempView("policy_summary")

        The documentation for the total_premium column might include:
        Derivation Table: "policy_summary"
        Derivation Column: "total_premium"
        Derivation Statements: "SELECT policy_id, SUM(premium_amount) as total_premium FROM policy_table WHERE status = 'Active' GROUP BY policy_id" or policy_table.filter("status = 'Active'").groupBy("policy_id").agg(sum("premium_amount").alias("total_premium"))
        Purpose of Derivation: "Calculate the total premium for each active policy using groupBy and aggregation."

        5. Identify User Defined Functions (UDFs) declarations (if any). Document the UDFs under "Table" along with any resulting variables from the UDF's logic under "Column". You will also document the join queries, derivation 
        logic and purpose for their derivation under the relevant columns.

        6. Since this may be a Spark script to provide data for a reporting dashboard, you will also analyse the section of the script which applies SELECT, UNION, JOIN and other such operations on the tables/DataFrames and columns 
        derived from the various stages of the transformation to procure the data for the report. These elements may not have table/DataFrame and column names, so you will assign a table/DataFrame and column name to them and an 
        accurate description when documenting this analysis. Suitable table/DataFrame names may be report, display_output, etc.

        7. Finally, you will populate the table thoroughly with the technical specifics and clear, contextual descriptions you will obtain from the above steps. The table should contain the following headers:
        | Section Name | Section Description | Table Name | Table Description | Column Name | Data Type | Column Description | Join Table | Join Column | Join Statements | Derivation Table | Derivation Column | Derivation Statements | Purpose of Derivation |
        |---|---|---|---|---|---|---|---|---|---|---|---|---|---|
        You will write N/A where values are not applicable. Ensure that all the rows begin and end with |. Truncate join statements and derivation statements to 8 keywords then add it to the documentation with "..." to denote 
        continuation. If a Section Name and consequently Section Description pertains to multiple tables/DataFrames and their columns, write S/B for subsequent instances. If a Table Name and consequently, Table Description pertains 
        to multiple columns, write S/B for repeating instances. If a Join Table, Join Column, Join Statement, Derivation Table, Derivation Column, Derivation Statement and consequently the Purpose of Derivation pertains to multiple 
        Column Names, fill the row for the first instance with the derivation statement as instructed above, and write "S/B" for subsequent instances.

        Following is the Spark SQL/PySpark script:
        {plsql_script}

        Following is the business glossary of the source system:
        {context}

        The resulting table will provide a detailed overview of the Spark script's operations, data flow, and business logic. It should serve as a reference guide for data engineers to understand the exact queries and operations applied, 
        and for business analysts to grasp the purpose and functionality of each section, element and statements within the script. Perform the documentation thoroughly and ensure no part of your analysis is excluded from the table as 
        this documentation is highly critical. Perform as much of the analysis as you can. Respond with as many complete rows of the table as you can create. if the table does not contain analysis and documentation of the entire Spark SQL 
        script, I will prompt you to continue and then you will continue populating the table. Do not include introductions, labels, backticks and conclusions. Only respond with the table. Each row of the table should end with a newline 
        character. If the analysis is complete for the entire stored procedure and the table is complete, do not start over from the beginning and end with ------. 
        """
        )

        continue_analysis_prompt = PromptTemplate.from_template(
            """ 
        You are a data engineer with expertise in understanding optimized, enterprise-grade Spark SQL and PySpark transformations with a focus on the BFSI sector. You were tasked with analyzing the provided Spark SQL/PySpark script which 
        is being used to build reports in the life insurance domain. You created part of a thorough and comprehensive table that captures the necessary details for both business analysts and data engineers by analysing the Spark SQL/PySpark 
        script and augmenting your analysis with the business glossary of the data source. The table includes information about multiple layers of transformations, tables/DataFrames and columns involved in each section, data types, key constraints, 
        description of each and every column, join tables/DataFrames and columns, join conditions, filter conditions, transformation statements, and the purpose of applying the joins, filters and transformations.

        Following is the table you have created so far:
        {table}

        Analyse the latest rows of the table created so far and determine where to analyse the Spark SQL/PySpark procedure from. Follow these instructions to continue creating the table:
        1. Analyse each and every section of the Spark script (Spark SQL or PySpark). The script consists of multiple layers of data transformation, designed to optimize query performance. Determine the purpose of each section in the data engineering 
        and report building process. You will utilize this analysis in naming and describing the sections as well as in documenting the tables/DataFrames involved in the section. Ensure that each section has only one concise and high-level description 
        documented under "Section Description".

        2. Identify the variables, tables/DataFrames, and columns in the Spark script. Analyse the select, from, and join statements, aggregations, CASE WHEN clauses, and other filter conditions. You will use this analysis to document the table/DataFrame 
        name, the columns belonging to the table/DataFrame and their data types in a hierarchical manner of data flow. You will also use this analysis to determine the purpose of the table/DataFrame and column in the report building process.

        3. Since this is a multi-stage transformation process, multiple tables/DataFrames and columns are derived with join statements, complex mathematical formulae, CASE WHEN conditions, aggregations, filters, etc. The derived tables/DataFrames and 
        columns may further be transformed into new tables/DataFrames and columns in a similar fashion. You will identify the statements pertaining to deriving a table/DataFrame and its column. You will analyse these statements and segment them into the 
        parts responsible for performing join operations and parts responsible for applying business logic such as filters based on CASE WHEN statements, aggregations, etc. If a statement does not contain any join operations, you will consider it to not 
        be applicable. You will map aliases to their actual table/DataFrame names and identify all the columns and tables/DataFrames being used in both sides of the various join operations, subsequent boolean operators and subqueries. You will identify 
        all the tables/DataFrames and columns being used in select, from, aggregation, CASE WHEN and other filter conditions for deriving a table/DataFrame and its columns. You will analyse these statements and use the analyses from steps 1 and 2 to comprehend 
        the derivation logic and the purpose of the derivation logic in the overall report building process. You will use these analyses to accurately document every table/DataFrame and column being invoked in the join part of the statements under join 
        table and join column separated by commas and you will only document the join operations in the statements under join statements. You will document the tables/DataFrames and columns invoked in CASE WHEN, aggregation and other filter conditions for 
        deriving a particular table/DataFrame and column under derivation table, derivation column, separated by commas, truncated versions of the derivation statements excluding the join statements, under derivation statement and the purpose and description 
        of the derivation logic being applied to derive them.
        For example, in the Spark SQL statement "INSERT INTO lins SELECT a.CHDrnum, CASE WHEN b.chdrnum IS NULL THEN 'O' ELSE 'R' END, a.ptrneff FROM PTRN1 a LEFT JOIN PTRNPF b ON a.chdrnum = b.chdrnum AND a.ptrneff = b.PTRNEFF WHERE b.validflag != '2' AND b.batctrcde = 'B522'", or its PySpark equivalent:
        lins = spark.sql("SELECT a.CHDrnum, CASE WHEN b.chdrnum IS NULL THEN 'O' ELSE 'R' END, a.ptrneff FROM PTRN1 a LEFT JOIN PTRNPF b ON a.chdrnum = b.chdrnum AND a.ptrneff = b.PTRNEFF WHERE b.validflag != '2' AND b.batctrcde = 'B522'")
        the join tables are "PTRN1, PTRNPF" which have been mapped to aliases 'a' and 'b' respectively, join columns are "PTRN1.chdrnum, PTRNPF.chdrnum, PTRN1.ptrneff, PTRNPF.PTRNEFF", 'PTRNPF.validflag' and 'PTRNPF.batctrcde' will not be included in the join columns 
        since they filter conditions, join statement is "FROM PTRN1 a LEFT JOIN PTRNPF b ON a.chdrnum = b.chdrnum AND a.ptrneff = b.PTRNEFF", derivation tables are "PTRN1, PTRNPF", derivation columns are "PTRNPF.chdrnum, PTRNPF.validflag, PTRNPF.batctrcde" and derivation 
        statement is "INSERT INTO lins SELECT a.CHDrnum, CASE WHEN b.chdrnum IS NULL THEN 'O' ELSE 'R' END...".

        4. If DataFrames are created using transformations like groupBy, agg, withColumn, document the DataFrame's name under "Derivation Table". Document the transformations used to create the DataFrame under "Derivation Statements," focusing on how they contribute 
        to the derivation of the table/DataFrame or column.
        For example, if a DataFrame policy_summary is created by grouping by policy_id and summing premium_amount:
        Spark SQL:
        CREATE OR REPLACE TEMP VIEW policy_summary AS
        SELECT policy_id, SUM(premium_amount) as total_premium FROM policy_table WHERE status = 'Active' GROUP BY policy_id

        PySpark:
        policy_summary = policy_table.filter("status = 'Active'").groupBy("policy_id").agg(sum("premium_amount").alias("total_premium"))
        policy_summary.createOrReplaceTempView("policy_summary")

        The documentation for the total_premium column might include:
        Derivation Table: "policy_summary"
        Derivation Column: "total_premium"
        Derivation Statements: "SELECT policy_id, SUM(premium_amount) as total_premium FROM policy_table WHERE status = 'Active' GROUP BY policy_id" or policy_table.filter("status = 'Active'").groupBy("policy_id").agg(sum("premium_amount").alias("total_premium"))
        Purpose of Derivation: "Calculate the total premium for each active policy using groupBy and aggregation."

        5. Identify User Defined Functions (UDFs) declarations (if any). Document the UDFs under "Table" along with any resulting variables from the UDF's logic under "Column". You will also document the join queries, derivation logic and purpose for their 
        derivation under the relevant columns.

        6. Since this may be a Spark script to provide data for a reporting dashboard, you will also analyse the section of the script which applies SELECT, UNION, JOIN and other such operations on the tables/DataFrames and columns derived from the various stages 
        of the transformation to procure the data for the report. These elements may not have table/DataFrame and column names, so you will assign a table/DataFrame and column name to them and an accurate description when documenting this analysis. Suitable table/DataFrame 
        names may be report, display_output, etc.

        7. Finally, you will populate the table thoroughly with the technical specifics and clear, contextual descriptions you will obtain from the above steps. The table should contain the following headers:
        | Section Name | Section Description | Table Name | Table Description | Column Name | Data Type | Column Description | Join Table | Join Column | Join Statements | Derivation Table | Derivation Column | Derivation Statements | Purpose of Derivation |
        |---|---|---|---|---|---|---|---|---|---|---|---|---|---|
        You will write N/A where values are not applicable. Ensure that all the rows begin and end with |. Truncate join statements and derivation statements to 8 keywords then add it to the documentation with "..." to denote continuation. If a Section Name and consequently 
        Section Description pertains to multiple tables/DataFrames and their columns, write S/B for repeating instances. If a Table Name and consequently, Table Description pertains to multiple columns, write S/B for repeating instances. If a Join Table, Join Column, Join Statement, 
        Derivation Table, Derivation Column, Derivation Statement and consequently the Purpose of Derivation pertains to multiple Column Names, fill the row for the first instance with the derivation statement as instructed above, and write "S/B" for subsequent instances.

        Following is the Spark SQL/PySpark script:
        {plsql_script}

        Following is the business glossary of the data source:
        {context}

        The resulting table will provide a detailed overview of the Spark script's operations, data flow, and business logic. It should serve as a reference guide for data engineers 
        to understand the exact queries and operations applied, and for business analysts to grasp the purpose and functionality of each section, element and statements within the script.
        Perform a complete analysis. Perform the documentation thoroughly. Do not include introductions, labels, backticks and conclusions. Dont include headers. Only respond with the new 
        rows of the table. Each row should end with a newline character. If the table is complete or has nearly been completed, i.e only one to three rows are remaining, respond with the remaining 
        rows if any, do not start over and end your response with ------.
        """
        )

    runnable = RunnablePassthrough()
    output_parser = StrOutputParser()

    chain = runnable | analysis_prompt | llm | output_parser
    continue_chain = runnable | continue_analysis_prompt | llm | output_parser

    table = chain.invoke({"plsql_script": plsql_script, "context": context})

    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(table)
    plsql_tokens = encoding.encode(plsql_script)

    while (
        "------" not in table.strip().split("\n")[-1]
        and len(tokens) + len(plsql_tokens) < 80000
    ):
        table = filter_table(table)
        continued_table = continue_chain.invoke(
            {"table": table, "plsql_script": plsql_script, "context": context}
        )
        continued_table = filter_table(continued_table)
        table += f"\n{continued_table}"

        if "---" in table.strip().split("\n")[-1]:
            break

        tokens = encoding.encode(table)

    df, df_filled = process_table(table)
    return df, df_filled


def visualise_etl(plsql_script, report_name, df):
    df = df.drop(
        columns=[
            "Description",
            "Table Description",
            "Data Type",
            "Column Description",
            "Join Statement",
            "Filter Statement",
        ]
    )
    documentation = df.to_string(index=False, header=True)
    diagram_prompt = PromptTemplate.from_template(
        """
    You will carefully analyse a detailed documentation of a stored procedure given below in a tabular format and accurately determine the table level data flow. You will then generate
    the graphviz source code for a data lineage diagram. You will create a single legend composed of a numbered list of the derivation purposes by analysing the entities involved in a 
    particular set of derivations detailed under "Derivation Purpose" in the documentation. For more context, "Table" refers to the entity being derived and "Join Table" as well as "Filter
    Table" document the entities being used to derive a particular "Table". "S/B" denotes that the value in that cell is the same as the value in the cell above. The edges between the nodes 
    of the entities should be accurately captioned with a number which will relate it to the numbered list of derivation purposes you are creating the legend for. After declaring the 
    nodes and edges, you will include labels and invisible edges within the same Digraph according to the "Layer" in the documentation and align the label of each layer with a logically 
    suitable entity derived in that particular layer. 

    Here's an example of how a node in the diagram should be formatted:
    PTRNPF [label="PTRNPF \n- chdrnum\n- ptrneff\n- batctrcde\n- validflag", fillcolor="#D0E8FF", color="#3B82F6"];
    PTRNPF may be a source table/temporary table/CTE/final report and chdrnum, ptrneff, batctrcde and validflag are its columns.

    Here are some styling guidelines for the nodes of entities:
    node shape : box
    node style : "rounded, filled"
    font : "Arial"

    source table fillcolor : #D0E8FF
    source table border color : #3B82F6
    temporary table/cursor/function fillcolor : #D1FAD7
    temporary table/cursor/function border color : #34D399
    CTE fillcolor : #FFEED6
    CTE border color : #FB923C
    final table/report/dbmsoutput fillcolor : #EDE7F6
    final table/report/dbmsoutput border color : #A855F7

    color of edges showing data flow to and from source tables, temporary tables/cursors and CTEs : #4B5563
    color of edges showing the data flow to the final report/dbmsoutput for data aggregation : #14B8A6
    style of edges showing filters only : dashed

    Here's an example of a Derivation Purpose legend:
        purposes_legend [label=<
            <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="0" CELLPADDING="3">
                <TR>
                    <TD BGCOLOR="white" ALIGN="CENTER"><FONT POINT-SIZE="14">Derivation Purpose</FONT></TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightblue" ALIGN="LEFT"><FONT POINT-SIZE="14">1. Filter policies based on batch code and effective date range</FONT></TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightblue" ALIGN="LEFT"><FONT POINT-SIZE="14">2. Determine payment flag based on the presence of policy in PTRNPF</FONT></TD>
                </TR>
            </TABLE>
        >]

    Here's an example of how you will format labels, declare invisibe edges and align layers:
    If the layers and the entities being derived in each layer are
    Initialization: N/A
    Cursor Declarations: BGMCUR, FLCCUR, BETCUR, DCEMCUR
    Function Definition: GETFXDCRNCYUNITS, GETFACILITYID
    Main Processing: BETCUR, BGMCUR, LLT, DUAL, LLTLC, GETFSXDCRNCYUNITS, DBMSOUTPUT

    label1 [label="Cursor Declarations", shape=box, style=filled, fillcolor=yellow, fontcolor=black, fontsize=12];
    label2 [label="Function Definition", shape=box, style=filled, fillcolor=yellow, fontcolor=black, fontsize=12];
    label3 [label="Main Processing", shape=box, style=filled, fillcolor=yellow, fontcolor=black, fontsize=12];

    edge[style=invis];
    label1 -> label2 -> label3

    {{ rank=same; label1; BGMCUR; BETCUR; FLCCUR; DCEMCUR }}
    {{ rank=same; label2; GETFXDCRNCYUNITS; GETFACILITYID }}
    {{ rank=same; label3; DBMSOUTPUT; LLT; LLTLC; DUAL }}

    Following are some general guidelines:
    - Including every unique source entity is important. However if the same set of operations is being performed within the stored procedure
    to derive the same set of entities across multiple layers is being derived separately across multiple layers you will declare the nodes for
    for those entities only once. Similarly, you will declare the edges to depict these derivations just once. The layers can logically 
    be considered to be a single layer. You will create a single label and assign it the name "Multiple Blocks".
    - All entity names should be complete and in upper case.
    - Remove the '@" prefix from the name of the entities, if present.
    - Layer labels must not include the entity name.
    - In the example of layers, "Initialization" is a variable initialisation layer, hence the respective entity provided is N/A. This layer 
    has been ignored and a label has not been declared for this layer.
    - In the example of layers, although entities like "BETCUR", "FACILITYID", etc are being derived in the "Main Processing" layer, they are 
    not being aligned with the "Main Processing" label since they have already been aligned with other layer labels.

    Following is the stored procedure:
    {plsql_script}

    Following is the documentation of the stored procedure in tabular format:
    {etl_documentation}

    Do not include introductions, explanations, backticks, labels or conclusions in your response. Respond with the complete graphviz source code 
    for the diagram.
    """
    )

    continue_diagram_prompt = PromptTemplate.from_template(
        """
    Earlier you had carefully analysed the stored procedure and its documentation provided to you and accurately determined the table level data flow. You utilised this analysis to generate 
    a portion of a graphviz source code for a lineage diagram. 

    Here are the instructions you followed to create generate the code:
    You will carefully analyse a detailed documentation of a stored procedure given below in a tabular format and accurately determine the table level data flow. You will then generate
    the graphviz source code for a data lineage diagram. You will create a single legend composed of a numbered list of the derivation purposes by analysing the entities involved in a 
    particular set of derivations detailed under "Derivation Purpose" in the documentation. For more context, "Table" refers to the entity being derived and "Join Table" as well as "Filter
    Table" document the entities being used to derive a particular "Table". "S/B" denotes that the value in that cell is the same as the value in the cell above. The edges between the nodes 
    of the entities should be accurately captioned with a number which will relate it to the numbered list of derivation purposes you are creating the legend for. After declaring the 
    nodes and edges, you will include labels and invisible edges within the same Digraph according to the "Layers" in the documentation and align the label of each layer with a logically suitable 
    entity derived in that particular layer. 

    Here's an example of how a node in the diagram should be formatted:
    PTRNPF [label="PTRNPF \n- chdrnum\n- ptrneff\n- batctrcde\n- validflag", fillcolor="#D0E8FF", color="#3B82F6"];
    PTRNPF may be a source table/temporary table/CTE/final report and chdrnum, ptrneff, batctrcde and validflag are its columns.

    Here are some styling guidelines for the nodes of entities:
    node shape : box
    node style : "rounded, filled"
    font : "Arial"

    source table fillcolor : #D0E8FF
    source table border color : #3B82F6
    temporary table/cursor/function fillcolor : #D1FAD7
    temporary table/cursor/function border color : #34D399
    CTE fillcolor : #FFEED6
    CTE border color : #FB923C
    final table/report/dbmsoutput fillcolor : #EDE7F6
    final table/report/dbmsoutput border color : #A855F7

    color of edges showing data flow to and from source tables, temporary tables/cursors and CTEs : #4B5563
    color of edges showing the data flow to the final report/dbmsoutput for data aggregation : #14B8A6
    style of edges showing filters only : dashed

    Here's an example of a Derivation Purpose legend:
        purposes_legend [label=<
            <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="0" CELLPADDING="3">
                <TR>
                    <TD BGCOLOR="white" ALIGN="CENTER"><FONT POINT-SIZE="14">Derivation Purpose</FONT></TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightblue" ALIGN="LEFT"><FONT POINT-SIZE="14">1. Filter policies based on batch code and effective date range</FONT></TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightblue" ALIGN="LEFT"><FONT POINT-SIZE="14">2. Determine payment flag based on the presence of policy in PTRNPF</FONT></TD>
                </TR>
            </TABLE>
        >]

    Here's an example of how you will format labels, declare invisibe edges and align layers:
    If the layers and the entities being derived in each layer are
    Initialization: N/A
    Cursor Declarations: BGMCUR, FLCCUR, BETCUR, DCEMCUR
    Function Definition: GETFXDCRNCYUNITS, GETFACILITYID
    Main Processing: BETCUR, BGMCUR, LLT, DUAL, LLTLC, GETFSXDCRNCYUNITS, DBMSOUTPUT

    label1 [label="Cursor Declarations", shape=box, style=filled, fillcolor=yellow, fontcolor=black, fontsize=12];
    label2 [label="Function Definition", shape=box, style=filled, fillcolor=yellow, fontcolor=black, fontsize=12];
    label3 [label="Main Processing", shape=box, style=filled, fillcolor=yellow, fontcolor=black, fontsize=12];

    edge[style=invis];
    label1 -> label2 -> label3

    {{ rank=same; label1; BGMCUR; BETCUR; FLCCUR; DCEMCUR }}
    {{ rank=same; label2; GETFXDCRNCYUNITS; GETFACILITYID }}
    {{ rank=same; label3; DBMSOUTPUT; LLT; LLTLC; DUAL }}

    Following are some general guidelines:
    - Including every unique source entity is important. However if the same set of operations is being performed within the stored procedure
    to derive the same set of entities across multiple layers is being derived separately across multiple layers you will declare the nodes for
    for those entities only once. Similarly, you will declare the edges to depict these derivations just once. The layers can logically 
    be considered to be a single layer. You will create a single label and assign it the name "Multiple Blocks".
    - All entity names should be complete and in upper case.
    - Remove the '@" prefix from the name of the entities, if present.
    - Layer labels must not include the entity name.
    - In the example of layers, "Initialization" is a variable initialisation layer, hence the respective entity provided is N/A. This layer 
    has been ignored and a label has not been declared for this layer.
    - In the example of layers, although entities like "BETCUR", "FACILITYID", etc are being derived in the "Main Processing" layer, they are 
    not being aligned with the "Main Processing" label since they have already been aligned with other layer labels.

    Following is the stored procedure:
    {plsql_script}

    Following is the documentation of the stored procedure in tabular format (rows from the top may have been removed for brevity):
    {etl_documentation}

    Following is the graphviz code you have generated so far (if you had started creating the Derivation Purpose legend then I will redact
    the earlier lines in the code for brevity; you will assume that the code for declaring the nodes, edges, labelling and alignment are already
    present and need not be written again):
    {graphviz_code}

    Do not include introductions, explanations, backticks, labels or conclusions in your response. Respond only with the remaining graphviz source
    code for the diagram. If the code for the diagram is close to completion, respond with the few remaining lines and ensure you end the code with }}.
    """
    )

    fix_code_prompt = PromptTemplate.from_template(
        """ 
    You are an expert graphviz coder. Your task is to debug the following code and fix syntax errors in it by understanding 
    its purpose and the error encountered on rendering a digram from a source object of this code:
    {code}

    Following was the error
    {error}

    Respond only with the complete fixed graphviz code. Do not include introductions, explanations and conclusions in your response.
    """
    )

    runnable = RunnablePassthrough()
    output_parser = StrOutputParser()

    diagram_chain = runnable | diagram_prompt | llm | output_parser
    continue_diagram_chain = runnable | continue_diagram_prompt | llm | output_parser
    fix_code_chain = runnable | fix_code_prompt | llm | output_parser

    graphviz_code = diagram_chain.invoke(
        {"plsql_script": plsql_script, "etl_documentation": documentation}
    )
    graphviz_code = (
        graphviz_code.replace("```plaintext", "")
        .replace("```graphviz", "")
        .replace("```dot", "")
        .strip("```")
        .strip()
    )

    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(graphviz_code)
    documentation_tokens = encoding.encode(documentation)

    lines_context = []
    graphviz_code_context = ""
    while (
        graphviz_code.rfind("}") != len(graphviz_code) - 1
        and len(tokens) + len(documentation_tokens) < 116800
    ):
        # while not graphviz_code.endswith("}\n}") and len(tokens) + len(documentation_tokens) < 116800:
        lines = graphviz_code.split("\n")
        if ">]" not in lines[-1]:
            lines.pop()
            while len(lines) > 0 and lines[-1] == lines[-2]:
                lines.pop()
            if lines_context == []:
                for i, line in enumerate(lines):
                    if "purposes_legend" in line:
                        lines_context = lines[i:]
        else:
            lines.append("}")
            graphviz_code = "\n".join(lines)
            break

        graphviz_code = "\n".join(lines)
        if lines_context != [] and graphviz_code_context == "":
            graphviz_code_context = "\n".join(lines_context)

        if 116800 - (len(tokens) + len(documentation_tokens)) < 6000:
            num_rows_remove = int(
                (126800 - (len(tokens) + len(documentation_tokens))) / 100
            )
            if df.shape[0] > 0 and df.shape[0] > num_rows_remove:
                df = df.iloc[num_rows_remove:]
                documentation = df.to_string(index=False, header=True)
                documentation_tokens = encoding.encode(documentation)
            else:
                continue

        continued_graphviz_code = continue_diagram_chain.invoke(
            {
                "plsql_script": plsql_script,
                "graphviz_code": (
                    graphviz_code_context
                    if graphviz_code_context != ""
                    else graphviz_code
                ),
                "etl_documentation": documentation,
            }
        )
        graphviz_code += f"\n{continued_graphviz_code}"
        if graphviz_code_context != "":
            graphviz_code_context = f"\n{continued_graphviz_code}"
        graphviz_code = (
            graphviz_code.replace("```plaintext", "")
            .replace("```graphviz", "")
            .replace("```dot", "")
            .strip("```")
            .strip()
        )
        graphviz_code_context = (
            graphviz_code_context.replace("```plaintext", "")
            .replace("```graphviz", "")
            .replace("```dot", "")
            .strip("```")
            .strip()
        )
        if graphviz_code_context != "":
            tokens = encoding.encode(graphviz_code_context)
        else:
            tokens = encoding.encode(graphviz_code)

    index = graphviz_code.find("{")
    legend = """
        node [shape=plaintext]
        legend [label=<
            <TABLE BORDER="0" CELLBORDER="1" CELLSPACING="0" CELLPADDING="4">
                <TR>
                    <TD COLSPAN="2"><B>Node Legend</B></TD>
                </TR>
                <TR>
                    <TD BGCOLOR="#D0E8FF"></TD>
                    <TD>Source Table</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="#D1FAD7"></TD>
                    <TD>Temp Tables, Cursors, Functions</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="#FFEED6"></TD>
                    <TD>CTE</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="#EDE7F6"></TD>
                    <TD>Final Table/Report</TD>
                </TR>
            </TABLE>
        >]

        edge_legend [label=<
            <TABLE BORDER="0" CELLBORDER="1" CELLSPACING="0" CELLPADDING="4">
                <TR>
                    <TD COLSPAN="2"><B>Edge Legend</B></TD>
                </TR>
                <TR>
                    <TD>
                        <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="0">
                            <TR>
                                <TD><FONT COLOR="#4B5563">______</FONT></TD>
                            </TR>
                        </TABLE>
                    </TD>
                    <TD>Data flow (source tables, temporary tables, CTEs)</TD>
                </TR>
                <TR>
                    <TD>
                        <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="0">
                            <TR>
                                <TD><FONT COLOR="#14B8A6">______</FONT></TD>
                            </TR>
                        </TABLE>
                    </TD>
                    <TD>Data flow to final report (aggregation)</TD>
                </TR>
                <TR>
                    <TD>
                        <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="0">
                            <TR>
                                <TD><FONT COLOR="#000000"><I>_ _ _</I></FONT></TD>
                            </TR>
                        </TABLE>
                    </TD>
                    <TD>Filters only (dashed)</TD>
                </TR>
            </TABLE>
        >]
    """
    graphviz_code = graphviz_code[: index + 1] + legend + graphviz_code[index + 1 :]

    LINEAGE_EXTRACTOR_OUTPUT_DIR = settings.FRONTEND_PUBLIC_DIR
    if not os.path.exists(LINEAGE_EXTRACTOR_OUTPUT_DIR):
        os.makedirs(LINEAGE_EXTRACTOR_OUTPUT_DIR)

    with open(
        f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{report_name.replace('.txt', '').replace('.sql', '')}_Lineage_Diagram_Code.txt",
        "w",
    ) as f:
        f.write(graphviz_code)

    image_path = f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{report_name.replace('.txt', '').replace('.sql', '')}_Lineage_Diagram"
    diagram = Source(graphviz_code, engine="dot")

    error = True
    i = 0
    while i < 5 and error:
        try:
            with open(
                f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{report_name.replace('.txt', '')}_Lineage_Diagram_Code.txt",
                "w",
            ) as f:
                f.write(graphviz_code)
            diagram = Source(graphviz_code, engine="dot")
            diagram.render(image_path, format="png", view=False, cleanup=True)
            image_path = f"{image_path}.png"
            return image_path
        except Exception as e:
            i += 1
            graphviz_code = fix_code_chain.invoke({"code": graphviz_code, "error": e})
            graphviz_code = (
                graphviz_code.replace("```plaintext", "")
                .replace("```graphviz", "")
                .replace("```dot", "")
                .strip("```")
                .strip()
            )
            continue


def create_etl_documentation_excel(df, report_name):
    workbook_name = report_name.replace(".txt", "_ETL_Documentation.xlsx").replace(
        ".sql", "_ETL_Documentation.xlsx"
    )
    sheet_name = "Analysis"
    LINEAGE_EXTRACTOR_OUTPUT_DIR = settings.FRONTEND_PUBLIC_DIR
    if not os.path.exists(LINEAGE_EXTRACTOR_OUTPUT_DIR):
        os.makedirs(LINEAGE_EXTRACTOR_OUTPUT_DIR)

    with pd.ExcelWriter(
        f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{workbook_name}", mode="w", engine="openpyxl"
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{workbook_name}")
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        start_row = 2
        for row in range(2, max_row + 1):
            current_value = ws.cell(row=row, column=col).value
            previous_value = ws.cell(row=row - 1, column=col).value

            if current_value == previous_value:
                continue
            else:
                if row - start_row > 1:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col,
                        end_row=row - 1,
                        end_column=col,
                    )
                start_row = row

        if max_row - start_row >= 1:
            ws.merge_cells(
                start_row=start_row, start_column=col, end_row=max_row, end_column=col
            )

    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    for cell in ws[1]:
        cell.fill = header_fill

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    for col in ws.iter_cols(min_col=1, max_col=max_col, min_row=2, max_row=max_row):
        col_letter = col[0].column_letter
        if col_letter == "J" or col_letter == "M":
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 15

    image_path = f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{report_name.replace('.txt', '').replace('.sql', '')}_Lineage_Diagram.png"
    if os.path.exists(image_path):
        ws = wb.create_sheet("Lineage Visualisation")
        image = Image(image_path)
        ws.add_image(image, "A1")

    wb.save(f"{LINEAGE_EXTRACTOR_OUTPUT_DIR}/{workbook_name}")


# Streamlit Interface
if "plsql_path" not in st.session_state:
    st.session_state["plsql_path"] = ""
if "report_name" not in st.session_state:
    st.session_state["report_name"] = None
if "plsql_script" not in st.session_state:
    st.session_state["plsql_script"] = None
if "plsql_dialect" not in st.session_state:
    st.session_state["plsql_dialect"] = None
if "context_for_plsql" not in st.session_state:
    st.session_state["context_for_plsql"] = ""
if "table_raw" not in st.session_state:
    st.session_state["table_raw"] = None
if "table" not in st.session_state:
    st.session_state["table"] = None
if "image_path" not in st.session_state:
    st.session_state["image_path"] = None

st.set_page_config(
    layout="wide",
    page_title="Automatic Data Lineage and Pipeline Analysis",
    page_icon="📝",
)

st.subheader(
    "Automatic Data Lineage and Pipeline Analysis of Stored Procedures alongwith Lineage Visualisation using AI 📝"
)
st.markdown(
    """
The accelerator supports teams by automating the analysis and documentation of Extractions, Transformations and Loading (ETL) operations
being performed within a Stored Procedure using Large Language Models. It accepts the path to a text file containing the source code of the 
stored procedure within it and optionally a business glossary of the source and target systems which enhances the Large Language Model's 
understanding of the business context.

The outcome is an Excel Workbook containing a well formatted tabular report which segments the various stages of the ETL process being performed in
the stored procedure, resolves every join operation and filter logic across multiple stages of transformation as well the tables involved in these 
operations, determines the business purpose of the transformations and describes the derived entities and attributes. The workbook contains a diagram
which visualises the lineage of the entities in the stored procedure. The workbook can be found in the **output/etl documentation** directory.

----
"""
)

plsql_path = st.text_input("Enter the path to the stored procedure:")
st.session_state["plsql_path"] = plsql_path.strip('"')
st.session_state["report_name"] = os.path.basename(st.session_state["plsql_path"])

if st.session_state["plsql_path"] != "":
    if (
        st.session_state["plsql_path"].split(".")[-1] == "txt"
        or st.session_state["plsql_path"].split(".")[-1] == "sql"
    ):
        for encoding in ["utf-8", "utf-16"]:
            try:
                with open(
                    st.session_state["plsql_path"], "r", encoding=encoding
                ) as plsql_script_file:
                    st.session_state["plsql_script"] = plsql_script_file.read()
                break
            except:
                continue
    else:
        st.info("Not a valid file")

plsql_dialect = st.selectbox("Choose dialect", ["T-SQL", "Spark SQL"])
st.session_state["plsql_dialect"] = plsql_dialect

context_for_plsql = st.text_area("Provide additional context for the stored procedure")
if context_for_plsql.strip() is not None:
    st.session_state["context_for_plsql"] += context_for_plsql.strip()

glossary_file = st.file_uploader("Upload a Business Glossary", type=["xlsx"])
if glossary_file and st.session_state["plsql_script"] is not None:
    st.session_state["context_for_plsql"] += "\n\nBusiness Glossary:\n" + get_glossary(
        st.session_state["plsql_script"], glossary_file
    )

if (
    st.session_state["plsql_script"] is not None
    and st.session_state["plsql_dialect"] is not None
):
    if st.button("Analyse"):
        with st.spinner("Analysing the stored procedure"):
            df, df_filled = analyse_sp(
                st.session_state["plsql_script"],
                st.session_state["context_for_plsql"],
                st.session_state["plsql_dialect"],
            )
            st.session_state["table"] = df_filled
            st.session_state["table_raw"] = df
            st.dataframe(st.session_state["table"])

if st.session_state["table_raw"] is not None:
    if st.button("Visualise Data Lineage"):
        with st.spinner("Generating Lineage Diagram"):
            try:
                image_path = visualise_etl(
                    st.session_state["plsql_script"],
                    st.session_state["report_name"],
                    st.session_state["table_raw"],
                )
                st.image(image_path, caption="Lineage Visualisation")
                st.session_state["image_path"] = image_path
            except Exception as e:
                st.error(f"An error occured! Please try again -> {e}")

    if st.button("Get Excel Sheet") and st.session_state["table"] is not None:
        with st.spinner("Creating Excel Sheet"):
            try:
                create_etl_documentation_excel(
                    st.session_state["table"], st.session_state["report_name"]
                )
                st.success("Excel sheet successfully saved in the local directory!")
            except Exception as e:
                st.error(
                    f"The documentation could not be generated! Please try again -> {e}"
                )
