import os
from typing import List

import pandas as pd

# Import LangChain components needed for parallel_description
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

from src.config.settings import settings

# Import centralized LLM and settings
from src.core.llm import (
    agent_llm,  # Assuming agent_llm is the one used for parallel_description
)

# Assuming embeddings are not directly used in these functions but rather by an external RAG component
# If embeddings are used here for RAG, they should also be imported from src.core.llm
# from langchain_community.embeddings import HuggingFaceEmbeddings


# --- Configuration Paths (now from settings) ---
# DATA_DICTIONARY_PATH = settings.DATA_DICTIONARY_PATH # Assuming you add this to settings
# EMBEDDINGS_PATH = settings.EMBEDDINGS_PATH # Assuming you add this to settings

# --- LLM Instance (now imported) ---
# gpt4o = agent_llm # Use the imported instance from src.core.llm.py

# --- Embeddings (if needed, import from core/llm.py or pass as arg) ---
# hf_embeddings = ... # (If actual RAG is done here, manage embeddings centrally too)


def read_schema_in_chunks_and_batch(
    schema_path: str, chunk_size: int = 1, batch_size: int = 1
) -> tuple[pd.DataFrame, list[list[str]]]:
    """
    Reads an Excel schema, processes it into batched chunks for LLM processing.

    Args:
        schema_path: Path to the schema Excel file.
        chunk_size: Number of table-column groups per chunk.
        batch_size: Number of chunks per batch.

    Returns:
        A tuple containing:
        - schema_df: The full DataFrame of the schema.
        - schema_info_batches: A list of batches, where each batch is a list of schema chunks (strings).
    """

    if not os.path.exists(schema_path):
        raise FileNotFoundError(f"Schema Excel file not found at '{schema_path}'")

    schema_df = pd.read_excel(schema_path, engine="openpyxl")

    schema_info_batches = []
    current_batch = []
    current_chunk = ""
    chunk_counter = 0

    for schema_name in schema_df["Schema Name"].unique():
        # Start a new chunk string for each schema to ensure logical grouping
        current_chunk = f"\nSCHEMA: {schema_name}\n"
        schema_tables = schema_df[schema_df["Schema Name"] == schema_name][
            "Table Name"
        ].unique()

        for table in schema_tables:
            # Check if adding this table would exceed the chunk_size or if it's a new batch cycle
            # This logic might need fine-tuning based on actual token limits of your LLM and data volume
            if (
                chunk_counter >= chunk_size
            ):  # Use >= to correctly trigger new chunk/batch
                current_batch.append(current_chunk)
                current_chunk = ""  # Reset current_chunk for the next group of tables
                chunk_counter = 0  # Reset chunk_counter

                if (
                    len(current_batch) >= batch_size
                ):  # Use >= to correctly trigger new batch
                    schema_info_batches.append(current_batch)
                    current_batch = []

            columns = (
                schema_df[
                    (schema_df["Schema Name"] == schema_name)
                    & (schema_df["Table Name"] == table)
                ]["Column Name"]
                .astype(str)
                .tolist()
            )

            # Breaking columns into sub-groups for readability within a table's chunk
            column_sub_chunks = []
            for i in range(
                0, len(columns), 12
            ):  # Adjust 12 as needed for column display per line
                column_sub_chunks.append(", ".join(columns[i : i + 12]))

            current_chunk += (
                f"\nTable: {table}\nColumns:\n" + ",\n".join(column_sub_chunks) + "\n"
            )
            chunk_counter += 1  # Increment for each table processed

        # After processing all tables for a schema, if there's an active chunk, add it to current_batch
        if current_chunk:
            current_batch.append(current_chunk)
            current_chunk = ""  # Clear for next schema
            chunk_counter = 0  # Reset for next schema
            if (
                len(current_batch) >= batch_size
            ):  # If current batch is full, add to main batches
                schema_info_batches.append(current_batch)
                current_batch = []

    # Add any remaining items in current_batch to schema_info_batches
    if current_batch:
        schema_info_batches.append(current_batch)

    return schema_df, schema_info_batches


# TODO: We have updated the function from asyncio.run to await. Can there be any implications?
async def parallel_description(
    schema_description: str,
    schema_info_batch: List[str],
    data_dictionary: List[str] = [],
) -> str:
    """
    Generates descriptions for schema information in parallel using an LLM.

    Args:
        schema_description: Overall description of the schema.
        schema_info_batch: A list of schema chunks (strings) to be processed in a batch.
        data_dictionary: Optional list of data dictionary entries for context.

    Returns:
        A concatenated string of all generated glossary parts.
    """

    description_prompt = PromptTemplate.from_template(
        """
    You are a business analyst and a subject matter expert in the life insurance domain. I will provide the business purpose and description of a schema
    alongwith schema info containing a list of columns and the tables they belong to. The schema description may also contain additional information about the naming
    conventions and other nomenclatures followed in the schema. You have to precisely describe each and every column so that a business glossary can be created.
    You will use your knowledge, augment it with your understanding of the information from the data dictionary, infer context and consider any additional information
    present in the schema description for describing the tables and columns by using the schema description and info as provided below:

    The data dictionary consists of data under the following fields in the given order, seperated by '->':
    Data Domain -> Sub Domain -> Field Name -> Field Description
    The Field Description not only contains description of the field but also other relevant information encapsulating business specifications.

    Data Dictionary: {data_dictionary}
    Schema Description: {schema_description}\n
    Schema Info: {schema_info}\n

    The response should be in the following format:
    Table: table name\n
    Description: description of the table\n
    - Column 1 name: column description\n
    - Column 2 name: column description\n
    ...

    Do not include introductions, explanations and conclusions. Ensure the column descriptions are accurate.
    """
    )

    requests = [
        {
            "schema_description": schema_description,
            "schema_info": schema_info_batch[i],  # Process each chunk in the batch
            "data_dictionary": (
                data_dictionary[i]
                if data_dictionary and i < len(data_dictionary)
                else "N/A"
            ),
        }
        for i in range(len(schema_info_batch))  # Iterate over the chunks in the batch
    ]

    runnable = RunnablePassthrough()
    output_parser = StrOutputParser()
    # Use agent_llm imported from src.core.llm
    describe_chain = runnable | description_prompt | agent_llm | output_parser

    # Use abatch for parallel asynchronous calls
    glossaries = await describe_chain.abatch(requests)

    if not isinstance(glossaries, list):
        raise TypeError("Expected glossaries to be a list of strings")

    glossary = "  \n".join(glossaries)
    return glossary


def create_business_glossary_excel(schema_name: str, glossary_content: str) -> str:
    """
    Creates an Excel file from the generated business glossary content.

    Args:
        schema_name: The base name for the output Excel file.
        glossary_content: The markdown/text content of the glossary.

    Returns:
        The path to the created Excel file.
    """

    business_glossary_file_path_name = (
        schema_name.replace(".xlsx", "") + "_With_Business_Glossary.xlsx"
    )
    output_full_path = os.path.join(
        settings.DEFINE_OUTPUT_DIR, business_glossary_file_path_name
    )
    sheet_name = "Business Glossary"

    lines = glossary_content.strip().split("\n")
    tables = []
    current_table = {}
    for line in lines:
        if line.startswith("Table:"):
            if current_table:
                tables.append(current_table)
            current_table = {
                "table_name": line.split("Table:")[1].strip(),
                "description": "",
                "columns": [],
            }
        elif line.startswith("Description:"):
            current_table["description"] = line.split("Description:")[1].strip()
        elif line.startswith("-"):
            column_detail = line.split(":", 1)
            if len(column_detail) > 1:
                column_name = column_detail[0].strip("-").strip()
                column_description = column_detail[1].strip()
                current_table["columns"].append((column_name, column_description))
            else:
                print(
                    f"Skipping improperly formatted line: {line.strip()}"
                )  # Use a logger here instead of print

    if current_table:
        tables.append(current_table)

    rows = []
    for table in tables:
        table_name = table["table_name"]
        table_description = table["description"]
        columns = table["columns"]
        for column_name, column_description in columns:
            rows.append(
                [table_name, table_description, column_name, column_description]
            )

    schema_df = pd.DataFrame(
        rows,
        columns=[
            "Table Name",
            "Table Description",
            "Column Name",
            "Column Description",
        ],
    )
    schema_df.to_excel(
        output_full_path,
        sheet_name=sheet_name,
        index=False,
    )

    # Excel formatting
    workbook = load_workbook(output_full_path)
    sheet = workbook[sheet_name]

    start_row = 2  # Assuming the first row is the header
    col_a_letter = "A"
    col_b_letter = "B"

    row = start_row
    while row <= sheet.max_row:
        cell_a = sheet[f"{col_a_letter}{row}"]
        cell_b = sheet[f"{col_b_letter}{row}"]

        if cell_a.value is not None:
            merge_start = row
            merge_end = row

            while (
                merge_end < sheet.max_row
                and sheet[f"{col_a_letter}{merge_end + 1}"].value == cell_a.value
            ):
                merge_end += 1

            if merge_end > merge_start:
                sheet.merge_cells(
                    start_row=merge_start,
                    start_column=cell_b.column,
                    end_row=merge_end,
                    end_column=cell_b.column,
                )
                merged_cell = sheet.cell(row=merge_start, column=cell_b.column)
                merged_cell.alignment = Alignment(wrapText=True, vertical="top")

            row = merge_end + 1
        else:
            row += 1

    header_fill = PatternFill(
        start_color="D0E8FA", end_color="D0E8FA", fill_type="solid"
    )
    for cell in sheet[1]:
        cell.fill = header_fill

    workbook.save(output_full_path)
    return output_full_path  # Return the path for the tool to report
