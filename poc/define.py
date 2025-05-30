import streamlit as st

import os
import regex as re
import pandas as pd
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from openpyxl import load_workbook, get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side

from langchain_openai import AzureChatOpenAI
# from langchain_ollama.llms import OllamaLLM 
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser


# Initialise LLM and RAG
gpt4o = AzureChatOpenAI(
    azure_deployment="gpt-4o",
    openai_api_version="2024-02-15-preview",
    api_key="901cce4a4b4045b39727bf1ce0e36219",
    azure_endpoint="https://aifordata-azureopenai.openai.azure.com/",
    temperature=0.5,
    max_tokens=4096
    )

# llama = OllamaLLM(model="llama3.3:70b-instruct-q8_0", base_url="http://host.docker.internal:11434")

def read_schema(schema_file=None, schema_df=None, batch_size=20):
    if schema_df is None:
        schema_df = pd.read_excel(schema_file, engine='openpyxl')
    schema_df = schema_df.drop_duplicates(subset=['Table Name', 'Column Name'], keep='first')
    schema_df = schema_df[schema_df['Table Name'].notna() & schema_df['Column Name'].notna()]
    schema_df = schema_df[schema_df['Schema Name'].notna() & schema_df['Table Name'].notna()]
    schema_df = schema_df[schema_df['Column Name'].str.strip() != '']
    schema_df = schema_df[schema_df['Table Name'].str.strip() != '']
    schema_df = schema_df[schema_df['Schema Name'].str.strip() != '']

    schema_info_batches = []
    current_batch = []
    has_column_name = 'Column Name' in schema_df.columns
    has_ddl = 'DDL' in schema_df.columns
    base_fields = {'Schema Name', 'Table Name', 'Column Name'}
    additional_fields = [col for col in schema_df.columns if col not in base_fields]    

    for schema in schema_df['Schema Name'].unique():
        schema_tables = schema_df[schema_df['Schema Name'] == schema]['Table Name'].unique()
        for table in schema_tables:
            chunk = f"\nSchema: {schema}\nTable: {table}\n"
            if has_ddl:
                ddl_val = schema_df[(schema_df['Schema Name'] == schema) & (schema_df['Table Name'] == table)]['DDL'].values[0]
                chunk += f"DDL: {ddl_val}\n"
            if has_column_name:
                # columns = [
                #     " ".join(col.split())
                #     for col in schema_df[
                #         (schema_df['Schema Name'] == schema) & (schema_df['Table Name'] == table)
                #     ]['Column Name'].astype(str).tolist()
                # ]
                # if columns:
                #     chunk += "Columns: " + ", ".join(columns) + "\n"
                table_rows = schema_df[(schema_df['Schema Name'] == schema) & (schema_df['Table Name'] == table)]
                columns = []
                for _, row in table_rows.iterrows():
                    col_name = str(row['Column Name']).strip()
                    # Gather additional field values for this row
                    extras = []
                    for field in additional_fields:
                        val = str(row[field]).strip() if pd.notna(row[field]) else ""
                        if val:
                            extras.append(val)
                    if extras:
                        col_name += " (" + ", ".join(extras) + ")"
                    columns.append(col_name)
                if columns:
                    chunk += "Columns: " + ", ".join(columns) + "\n"
            current_batch.append(chunk)
            if len(current_batch) == batch_size:
                schema_info_batches.append(current_batch)
                current_batch = []
    if current_batch:
        schema_info_batches.append(current_batch)

    return schema_df, schema_info_batches

def ddl_description(chunk, schema_description):
    db_schema_prompt = PromptTemplate.from_template(
    """
    You are a business analyst and subject matter expertise in multiple domains of enterprise including ESG, BFSI and retail. 
    You will create a comprehensive business glossary from the given database information by reasoning step-by-step 
    based on the following instruction:

    1. Table Description: For each table, generate an informative description by drawing context from the schema description as well as name 
    and metadata of the attributes. Be authoritative and concise, using your expertise in the domain to provide a clear understanding of the 
    table's purpose and relevance. Table descriptions should be succinct yet informative, capturing the essence of the data it holds and its role 
    in the broader business context. Do not use generic terms or broad and vague descriptions; do not use words like "likely", "may", 
    "appears to", "could", etc. Instead, use definitive language that reflects your expertise and understanding of the domain. 

    2. Column Descriptions: List all columns for each table with comprehensive and business-centric definitions. Industry specific terminologies
    and business concepts should be elaborated upon such that the reader finds the glossary useful for data engineering as well as business analytics.
    Abbreviations should be expanded and explained, and acronyms should be defined. Do not use generic terms or broad and vague descriptions.
    Ensure there is no redundancy in the descriptions. The description should be accompanied by the subject area or subdomain under the business domain
    that the data within the column pertains to. Subdomain name should be enclosed in brackets.

    3. Infer meanings for incomplete or ambiguous names based on context from the metadata. 

    4. Only describe the tables and columns listed below. Do not invent or add any that are not present.

    Schema Description:
    {schema_description}
    
    Table Info: 
    {schema_data}

    Use this format throughout:
    Schema: [Schema Name]\n
    Table: [Table Name]\n
    Description: [Concise description of the table]\n                            
    - [Column 1 Name]: [Comprehensive and business-centric definition for the column] ([Subdomain])\n  
    - [Column 2 Name]: [Comprehensive and business-centric definition for the column] ([Subdomain])\n 
    ...
    - [Column n Name]: [Comprehensive and business-centric definition for the column] ([Subdomain])
    -------
                                                    
    Do not include introductory or closing messages. After describing all columns, end your response with ------.
    """)

    db_schema_continue_prompt = PromptTemplate.from_template(
    """
    You are a business analyst and subject matter expertise in multiple domains of enterprise including ESG, BFSI and retail. 
    You created part of a comprehensive business glossary from the given database information by reasoning step-by-step 
    based on the following instruction:

    1. Table Description: For each table, generate an informative description by drawing context from the schema description as well as name 
    and metadata of the attributes. Be authoritative and concise, using your expertise in the domain to provide a clear understanding of the 
    table's purpose and relevance. Table descriptions should be succinct yet informative, capturing the essence of the data it holds and its role 
    in the broader business context. Do not use generic terms or broad and vague descriptions; do not use words like "likely", "may", 
    "appears to", "could", etc. Instead, use definitive language that reflects your expertise and understanding of the domain. 

    2. Column Descriptions: List all columns for each table with comprehensive and business-centric definitions. Industry specific terminologies
    should be elaborated upon such that the reader finds the glossary useful for data engineering as well as business analytics.
    Abbreviations should be expanded and explained, and acronyms should be defined. Do not use generic terms or broad and vague descriptions.
    Ensure there is no redundancy in the descriptions. The description should be accompanied by the subject area or subdomain under the business domain
    that the data within the column pertains to. Subdomain name should be enclosed in brackets.

    3. Infer meanings for incomplete or ambiguous names based on context from the metadata. 

    4. Only describe the tables and columns listed below. Do not invent or add any that are not present.

    Schema Description:
    {schema_description}

    Table Info: 
    {schema_data}

    Glossary generated so far:
    {glossary}

    Continue from the last point and complete the glossary by providing definitions for the remaining columns.

    Use this format throughout:                          
    - [Remaining Column 1 Name]: [Comprehensive and business-centric definition for the column] ([Subdomain])\n  
    - [Remaining Column 2 Name]: [Comprehensive and business-centric definition for the column] ([Subdomain])\n 
    ...
    - [Remaining Column n Name]: [Comprehensive and business-centric definition for the column] ([Subdomain]) 
    ------
                                                    
    Do not include introductory or closing messages. After describing all columns, end your response with ------.
    """)

    runnable = RunnablePassthrough()
    output_parser = StrOutputParser()

    chain = runnable | db_schema_prompt | gpt4o | output_parser
    continue_chain = runnable | db_schema_continue_prompt | gpt4o | output_parser

    glossary = chain.invoke({"schema_data": chunk, "schema_description": schema_description})

    while not glossary.endswith("------"):
        glossary_lines = glossary.split('\n')
        if chunk.split(", ")[-1].split(" (")[0] != glossary_lines[-1].split(": ")[0].strip("- "):
            glossary = '\n'.join(glossary_lines[:-1])
        else:
            glossary_lines[-1] = glossary_lines[-1].split(": ")[0] + "Could not be generated.\n------"
            glossary = '\n'.join(glossary_lines)
            break
        continued_glossary = continue_chain.invoke({"schema_data": chunk, "glossary": glossary, "schema_description": schema_description})
        glossary += f"\n{continued_glossary}"

    return glossary

def generate_glossary_ddl(schema_info, schema_description):
    glossaries = []
    futures = []
    with ThreadPoolExecutor(max_workers=20) as executor:
        for chunk in schema_info:
            future = executor.submit(ddl_description, chunk, schema_description)
            futures.append(future)

    glossaries.extend([future.result() for future in futures])

    glossary = "  \n".join(glossaries)
    return glossary
    
def create_business_glossary_dataframe(glossary):
    lines = glossary.strip().split('\n')
    tables = []
    current_schema = ""
    current_table = {}
    for line in lines:
        if line.startswith('Schema:'):
            current_schema = line.split('Schema:')[1].strip()
        elif line.startswith('Table:'):
            if current_table:
                tables.append(current_table)
            current_table = {
                'schema_name': current_schema,
                'table_name': line.split('Table:')[1].strip(),
                'description': '',
                'columns': []
            }
        elif line.startswith('Description:'):
            current_table['description'] = line.split('Description:')[1].strip()
        elif line.startswith('-'):
            column_detail = line.split(':', 1)
            if len(column_detail) > 1:
                column_name = column_detail[0].strip('-').strip()
                desc = column_detail[1].strip()
                # Extract description and subdomain using regex
                match = re.match(r'^(.*)\s+\(([^()]*)\)\s*$', desc)
                if match:
                    column_description = match.group(1).strip()
                    column_subdomain = match.group(2).strip()
                else:
                    column_description = desc
                    column_subdomain = ""
                current_table['columns'].append((column_name, column_description, column_subdomain))
            else:
                print(f"Skipping improperly formatted line: {line.strip()}")
    if current_table:
        tables.append(current_table)

    optional_fields = [field for field in st.session_state['schema_df'].columns if field not in ['Schema Name', 'Table Name', 'Column Name']]

    rows = []
    for table in tables:
        schema_name = table['schema_name']
        table_name = table['table_name']
        table_description = table['description']
        columns = table['columns']
        for column_name, column_description, column_subdomain in columns:
            extra_values = []
            for field in optional_fields:
                val = ""
                match_row = schema_df[
                    (schema_df['Schema Name'] == schema_name) &
                    (schema_df['Table Name'] == table_name) &
                    (schema_df['Column Name'] == column_name)
                ]
                if not match_row.empty:
                    val = str(match_row.iloc[0][field])
                extra_values.append(val)
            row = [schema_name, table_name, table_description, column_name, column_description, column_subdomain] + extra_values
            rows.append(row)
    
    columns_list = ['Schema Name', 'Table Name', 'Table Description', 'Column Name', 'Column Description', 'Column Subdomain'] + optional_fields
    glossary_df = pd.DataFrame(rows, columns=columns_list)

    glossary_df = glossary_df.drop_duplicates(subset=['Schema Name', 'Table Name', 'Column Name'], keep='first')

    valid_tables = set(zip(
        st.session_state['schema_df']['Schema Name'], 
        st.session_state['schema_df']['Table Name'], 
        st.session_state['schema_df']['Column Name']))
    filtered_rows = [row for row in rows if (row[0], row[1], row[3]) in valid_tables]
    glossary_df = pd.DataFrame(filtered_rows, columns=columns_list)
    
    original_pairs = set(zip(
        st.session_state['schema_df']['Schema Name'], 
        st.session_state['schema_df']['Table Name'], 
        st.session_state['schema_df']['Column Name']))
    described_pairs = set(zip(
        glossary_df['Schema Name'], 
        glossary_df['Table Name'], 
        glossary_df['Column Name']))
    missing_pairs = original_pairs - described_pairs

    if missing_pairs:
        missing_df = pd.DataFrame(list(missing_pairs), columns=['Schema Name', 'Table Name', 'Column Name'])
        return glossary_df, missing_df
    else:
        return glossary_df, None

def create_business_glossary_excel(schema_name, schema_df):
    business_glossary_file_path = schema_name.replace('.xlsx', '') + "_with_Business_Glossary.xlsx"
    sheet_name = "Business Glossary"

    if not os.path.exists(f"./output/business glossary"):
        os.makedirs(f"./output/business glossary")
    schema_df.to_excel(f"./output/business glossary/{business_glossary_file_path}", sheet_name=sheet_name, index=False)

    workbook = load_workbook(f"./output/business glossary/{business_glossary_file_path}")
    sheet = workbook[sheet_name]

    # Set column widths dynamically, but keep your preferred widths for known columns
    col_widths = {
        "Schema Name": 26,
        "Table Name": 30,
        "Table Description": 26,
        "Column Name": 30,
        "Column Description": 26,
        "Column Subdomain": 26,
    }
    for idx, col_name in enumerate(schema_df.columns, 1):
        width = col_widths.get(col_name, 20)
        sheet.column_dimensions[get_column_letter(idx)].width = width

    # Freeze top row
    sheet.freeze_panes = sheet['A2']

    # Define border style
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set alignment and borders for all cells, and set row height for header
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for idx, cell in enumerate(row):
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = border
            if cell.row == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if row[0].row == 1:
            sheet.row_dimensions[1].height = 30
        else:
            # Left justify all except Table Name (center)
            for i, c in enumerate(row):
                col_name = schema_df.columns[i]
                if col_name == "Table Name":
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Merge Table Description cells for the same table
    # Find the column index for "Table Description"
    try:
        table_desc_idx = list(schema_df.columns).index("Table Description") + 1  # 1-based for openpyxl
        table_name_idx = list(schema_df.columns).index("Table Name") + 1
    except ValueError:
        table_desc_idx = 3  # fallback to column C
        table_name_idx = 2  # fallback to column B

    start_row = 2
    row = start_row
    while row <= sheet.max_row:
        cell_table_name = sheet.cell(row=row, column=table_name_idx)
        cell_table_desc = sheet.cell(row=row, column=table_desc_idx)
        if cell_table_name.value is not None:
            merge_start = row
            merge_end = row
            # Merge for all consecutive rows with the same Table Name
            while merge_end < sheet.max_row and sheet.cell(row=merge_end + 1, column=table_name_idx).value == cell_table_name.value:
                merge_end += 1
            if merge_end > merge_start:
                sheet.merge_cells(start_row=merge_start, start_column=table_desc_idx, end_row=merge_end, end_column=table_desc_idx)
                merged_cell = sheet.cell(row=merge_start, column=table_desc_idx)
                merged_cell.alignment = Alignment(wrap_text=True, vertical='top')
            row = merge_end + 1
        else:
            row += 1

    # Header fill
    header_fill = PatternFill(start_color="D0E8FA", end_color="D0E8FA", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = header_fill

    workbook.save(f"./output/business glossary/{business_glossary_file_path}")
    return f"./output/business glossary/{business_glossary_file_path}"


# Streamlit Interface
if 'schema_info' not in st.session_state:
    st.session_state['schema_info'] = ""
if 'schema_info_batches' not in st.session_state:
    st.session_state['schema_info_batches'] = []
if 'schema_df' not in st.session_state:  
    st.session_state['schema_df'] = None
if 'glossary' not in st.session_state:
    st.session_state['glossary'] = ""
if 'glossary_continuation' not in st.session_state:
    st.session_state['glossary_continuation'] = ""
if 'done' not in st.session_state:
    st.session_state['done'] = False
if 'missing_schema_df' not in st.session_state:
    st.session_state['missing_schema_df'] = None
if 'glossary_df' not in st.session_state:
    st.session_state['glossary_df'] = None


st.set_page_config(
    layout="wide",
    page_title="Automatic Generation Business Glossary",
    page_icon="📖",
)

st.subheader("Automatic Creation of Business Glossary using EY's Expertise and Large Language Models 📖")
st.markdown(
"""
The accelerator supports teams in cataloging source systems which may often contain ambiguous names for entities and attributes. 
It resolves such ambiguities by leveraging EY's extensive experience in the financial and life insurance sector. It accepts the 
path to any excel sheet containing a list of table names and column names and a brief description of the source system where you
may include further specifications regarding updates and naming conventions. The accelerator relies on a RAG pipeline to a knowledge 
base and a Large Language Model.

The outcome is an excel sheet containing table and column descriptions, which can be found in the **output/business glossary** directory.

----
""")

schema_file = st.file_uploader("Upload Schema File (Should contain fields: Schema Name, Table Name and Column Name)", type=["xlsx"])
schema_description = st.text_area("Enter the schema description:")

if schema_file:
    schema_df, schema_info_batches = read_schema(schema_file=schema_file)
    st.session_state['schema_info_batches'] = schema_info_batches
    st.session_state['schema_df'] = schema_df

if schema_description and st.session_state['schema_info_batches'] != []:
    schema_name = schema_file.name

    st.subheader("Tables and Columns", divider='violet')
    for batch in st.session_state['schema_info_batches']:
        show_schema = ''.join(batch)
        st.text(show_schema)

    if st.session_state['glossary'] != "":
        if st.session_state['missing_schema_df'] is None:
            glossary_df, missing_schema_df = create_business_glossary_dataframe(st.session_state['glossary'])
            st.session_state['glossary_df'] = glossary_df
            st.session_state['missing_schema_df'] = missing_schema_df
            if st.session_state['missing_schema_df'] is None:
                st.session_state['schema_info_batches'] = None
        st.subheader("Business Glossary", divider='violet')
        st.markdown(st.session_state['glossary'])

    if st.session_state['missing_schema_df'] is not None:
        st.subheader("Business Glossary yet to be generated for", divider='violet')
        st.write(f"{len(st.session_state['missing_schema_df'])} table/column pairs from the schema were not described.")
        st.dataframe(st.session_state['missing_schema_df'])
        st.write("Generate the remaining glossary by clicking the Describe button.")
        _, schema_info_batches = read_schema(schema_df=st.session_state['missing_schema_df'])
        st.session_state['schema_info_batches'] = schema_info_batches

    if st.session_state['schema_info_batches'] is not None:
        if st.button("Describe"):
            for batch_index, batch in enumerate(st.session_state['schema_info_batches']):
                try:
                    with st.spinner(f"Generating Business Glossary for batch {batch_index + 1} of {len(st.session_state['schema_info_batches'])}"):
                        glossary = generate_glossary_ddl(batch, schema_description)
                        st.session_state['glossary_continuation'] = f"  \n{glossary}"
                        st.session_state['glossary'] += st.session_state['glossary_continuation']
                    st.markdown(st.session_state['glossary_continuation'])
                except Exception as e:
                    print(f"Exception at batch {batch_index + 1}: ", e)
                    continue
                if batch_index + 1 == len(st.session_state['schema_info_batches']):
                    st.session_state['done'] = True
    
    if st.session_state['done']:
        st.session_state['done'] = False
        st.rerun()
    
    # Button to get Business Glossary Excel sheet
    if st.session_state['glossary_df'] is not None:
        if st.button("Create Business Glossary Excel Sheet"):
            file_path = create_business_glossary_excel(schema_name, st.session_state['glossary_df'])
            if st.session_state['missing_schema_df'] is not None:
                st.session_state['missing_schema_df'].to_excel(file_path, sheet_name="Unknown Attributes", index=False)
            st.success("Business Glossary Excel sheet ready for download! Please proceed")
            st.session_state['glossary'] = ""
            st.session_state['glossary_continuation'] = ""
            with open(file_path, "rb") as f:
                st.download_button(
                    label="Download Business Glossary Excel Sheet",
                    data=f,
                    file_name=os.path.basename(file_path),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=lambda: os.remove(file_path)
                )
