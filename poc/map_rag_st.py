import streamlit as st
import pandas as pd
import json
import numpy as np
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
import pickle
import tiktoken

from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

from langchain_openai import AzureChatOpenAI
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser

from graphviz import Digraph

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


llm = AzureChatOpenAI(
    azure_deployment="gpt-4o",
    openai_api_version="2024-02-15-preview",
    api_key="901cce4a4b4045b39727bf1ce0e36219",
    azure_endpoint="https://aifordata-azureopenai.openai.azure.com/",
    temperature=0,
    max_tokens=4096
)

hf_embeddings = SentenceTransformer("./all-mpnet-base-v2/", trust_remote_code=True)
encoding = tiktoken.get_encoding("cl100k_base")

runnable = RunnablePassthrough()
output_parser = StrOutputParser()

metadata_list = []
table_data_list = []

def create_embeddings(row):
    metadata = {
        "table": str(row['Table Name']),
        "table_description": str(row['Table Description']),
        "column": str(row['Column Name']),
        "column_description": str(row['Column Description'])
    }
    column_description_embedding = hf_embeddings.encode(metadata["column_description"])
    metadata["column_description_embedding"] = column_description_embedding.tolist()

    return metadata

def create_table_embeddings(row):
    table_data = {
        "table": str(row['Table Name']),
    }
    table_description_embedding = hf_embeddings.encode(str(row['Table Description']))
    table_data["table_description_embedding"] = table_description_embedding.tolist()  
        
    return table_data

def create_index(file):
    try:
        df = pd.read_excel(file)
        df.ffill(inplace=True)
    except Exception as e:
        print(e)

    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = []
        for _, row in df.iterrows():
            future = executor.submit(create_embeddings, row)
            futures.append(future)
    
    for future in futures:
        metadata_list.append(future.result())

    table_df = df.groupby(['Table Name'])[['Table Name', 'Table Description']].first()

    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = []
        for _, row in table_df.iterrows():
            future = executor.submit(create_table_embeddings, row)
            futures.append(future)
    
    for future in futures:
        table_data_list.append(future.result())

    with open(f"./input/{file.name.replace('.xlsx', '')}_metadata.pkl", 'wb') as f:
        pickle.dump(metadata_list, f)

    with open(f"./input/{file.name.replace('.xlsx', '')}_table_data.pkl", 'wb') as f:
        pickle.dump(table_data_list, f)

    return metadata_list, table_data_list

def search_glossary_entity(query, top_k=100):
    query_embedding = hf_embeddings.encode(query)
    table_list = []
    similarity_score_list = []
    results = []
    for metadata in table_data_list:
        stored_embeddings = metadata.get("table_description_embedding")
        table = metadata.get("table")
        if table not in table_list:
            table_list.append(table)
            stored_embedding = np.array(stored_embeddings)
            similarity_score = cosine_similarity([query_embedding], [stored_embedding])[0][0]
            results.append((table, similarity_score))
            similarity_score_list.append(similarity_score)
    
    # upper_quartile = np.percentile(similarity_score_list, 75)
    
    if len(results) > 0:
        results.sort(key=lambda item: item[1], reverse=True)
        # return [table for table, value in results[:top_k] if value >= upper_quartile]
        return [table for table, _ in results[:top_k]]

    return None

def search_glossary_attribute(query, source_entities, top_k=5):
    query_embedding = hf_embeddings.encode(query)
    similarity_score_list = []
    results = []
    for metadata in [md for md in metadata_list if md.get('table') in source_entities]:
        stored_embeddings = metadata.get("column_description_embedding")
        if stored_embeddings:
            stored_embedding = np.array(stored_embeddings)
            similarity_score = cosine_similarity([query_embedding], [stored_embedding])[0][0]
            results.append((metadata, similarity_score))
            similarity_score_list.append(similarity_score)

    upper_quartile = np.percentile(similarity_score_list, 75)

    if len(results) > 0:
        results.sort(key=lambda item: item[1], reverse=True)
        return '\n'.join([f"{result['table']} | {result['table_description']} | {result['column']} | {result['column_description']}" for result, value in results[:top_k] if value >= upper_quartile])

    return None

def search_glossary_pk(source_entity, pk_description, top_k=3):
    query_embedding = hf_embeddings.encode(pk_description)
    similarity_score_list = []
    results = []
    for metadata in [md for md in metadata_list if md.get('table') == source_entity]:
        stored_embeddings = metadata.get("column_description_embedding")
        if stored_embeddings:
            stored_embedding = np.array(stored_embeddings)
            similarity_score = cosine_similarity([query_embedding], [stored_embedding])[0][0]
            results.append((metadata, similarity_score))
            similarity_score_list.append(similarity_score)
        
    upper_quartile = np.percentile(similarity_score_list, 95)

    if len(results) > 0:
        results.sort(key=lambda item: item[1], reverse=True)
        return '\n'.join([f"{result['table']} | {result['table_description']} | {result['column']} | {result['column_description']}" for result, value in results[:top_k] if value >= upper_quartile])

    return None

def process_future_result(future, pk_name, pk_description):
    # source_attributes, attribute = future.result()
    attribute, source_attributes = future.result()
    # pk_dict = {}
    # row_values = []
    row_values_list = []

    if source_attributes is not None:
        for source_attribute in source_attributes.split('\n'):
            result_elements = source_attribute.split(' | ')
            pk_source_entities = search_glossary_pk(result_elements[0], pk_description)
            if pk_source_entities is not None:
                for pk_source_entity in pk_source_entities.split('\n'):
                    if all(pk_word in pk_source_entity.split(' | ')[3] for pk_word in pk_name.split()) and (
                        "identifier" in pk_source_entity.split(' | ')[3].lower() 
                        or "unique" in pk_source_entity.split(' | ')[3].lower()
                    ):
                        row_values = [
                            "Silver",
                            result_elements[0],
                            result_elements[1], 
                            result_elements[2], 
                            result_elements[3], 
                            pk_source_entity.split(' | ')[2], 
                            pk_source_entity.split(' | ')[3], 
                            attribute['Entity_Name'], 
                            attribute['Entity_Description'],
                            attribute['Attribute_Name'],
                            attribute['Attribute_Description'],
                            attribute['Data_Type'],
                            ""    
                        ]
                        row_values_list.append(row_values)
                        break
                              
    return row_values_list

mapping_prompt = PromptTemplate.from_template(
""" 
You are a highly experienced analyst in the life insurance domain, specializing in data integration and mapping for analytical purposes. 
Your objective is to meticulously analyze a provided mapping table and create *only* accurate and sensible mapping rules. You will be 
given:

* **Mapping Table:**  Details source and target entity/attribute mappings, along with source key constraints and target data types.
* **Glossary:**  Comprehensive descriptions of all source and target entities and attributes.

Your task is to analyze each mapping row and determine if it represents an accurate and business-sensible mapping.  
**Crucially, identify and exclude inaccurate or illogical mappings.** For accurate mappings, define the appropriate mapping rule.

**Crucial Definitions:**

* **Accurate Mapping:** A mapping that is not only data type compatible but also semantically consistent and business-relevant. It must 
make logical sense within the context of life insurance operations and analytics.
* **Direct Mapping:** A one-to-one mapping where the source attribute can be directly transferred to the target attribute *without any 
transformation or business logic*. This implies a strong semantic similarity and purpose alignment between the source and target attributes 
as defined in the glossary. **You must explicitly justify why a mapping is deemed "Direct" based on glossary descriptions.**
* **Logic/Transformation Required:**  The mapping is accurate in principle, but requires business logic, calculations, aggregations, data 
type conversions, or combinations of source attributes to derive the target attribute.

**Guidelines for Analysis (Follow these steps meticulously for each mapping row):**

1. **Deep Glossary Dive:**  For each source entity/attribute and target entity/attribute in the mapping row, **carefully read and understand
their descriptions in the provided glossary.**  Pay close attention to the stated purpose, business context, and any relationships described.

2. **Semantic Alignment Check:** Based on the glossary descriptions and your life insurance domain knowledge, evaluate:
    * **Does the source attribute semantically represent the same concept as the target attribute?**  Are they measuring or describing the 
    same business element?
    * **Is there a clear and logical business reason for mapping the source to the target?**  Consider the analytical purpose of the target 
    data model and the transactional nature of the source system.
    * **Are there any semantic discrepancies or mismatches based on the glossary definitions?**

3. **Logic/Transformation Necessity Assessment:**
    * **Direct Mapping Potential?**  If the semantic alignment is exceptionally strong after step 2, consider if a "Direct" mapping is truly
    justified. **You must justify this "Direct" assessment based on the glossary descriptions.**
    * **Logic Required?** If there are any semantic nuances, data type differences, or if deriving the target attribute from the source 
    attribute requires any business rule or calculation within the life insurance context, classify it as requiring "Logic/Transformation."  Consider scenarios like:
        * **Calculations or Aggregations:**  Source attributes might represent components that need to be combined to derive the target 
        attribute (e.g., combining premium components).
        * **Conditional Logic:** The mapping might depend on specific conditions or statuses within the source data.
        * **Data Type Conversions with Business Implications:**  Simple data type conversions might still require business context understanding.

4. **Inaccurate Mapping Identification:**
    * **Semantic Mismatch:** If, after steps 1-3, you identify a clear semantic mismatch or illogical business connection between the 
    source and target attributes based on the glossary, **classify the mapping as inaccurate and exclude it from your output.**
    * **Data Type Incompatibility without Sensible Transformation:** If data types are fundamentally incompatible and no reasonable 
    business transformation can bridge the gap, classify it as inaccurate.

5. **Mapping Rule Formulation:**
    * **Direct Mapping:** If justified as "Direct," the Mapping Rule is "Direct."
    * **Logic/Transformation Required:**  Describe the necessary business logic or transformation concisely but clearly in the 
    "Mapping Rule" column. Be specific about the operations needed (e.g., "Sum of SourceAttribute1 and SourceAttribute2," "Convert data 
    type to Date and format as YYYY-MM-DD," "Use lookup table based on SourceAttribute3").

**Output Format:**

For each *accurate* mapping you identify, provide a row in the following table format:

|[Source Entity Name]|[Source Attribute Name]|[Target Entity Name]|[Target Attribute Name]|[Mapping Rule]|
...
------

**Remember**: Only output rows for *accurate* and *business-sensible* mappings.  Justify "Direct" mappings and clearly describe 
"Logic/Transformation" rules. Do not include introductions, explanations, or conclusions. Complete the table and end your response 
with ------.

{context}
"""
)

sql_prompt = PromptTemplate.from_template(
""" 
You are a data engineer with expertise in writing optimised, enterprise-grade queries to perform transformations and mappings alongwith subject matter
expertise in the life insurance domain. Your task is to analyse the mapping table which will be provided to you and to write the accurate SQL query 
to perform the mappings by analysing the mapping rules.

Following is the mapping table:
{context}

SQL:
```sql
[mapping query]
```

Respond only with the complete SQL query. Do not include introductions, explanations and conclusions.
"""
)

mapping_chain = runnable | mapping_prompt | llm | output_parser
sql_chain = runnable | sql_prompt | llm | output_parser

columns_to_check = ['Source Entity', 'Source Attribute', 'Target Entity', 'Target Attribute']

def process_mapping_df(initial_mapping_df, pk):
    entity_descriptions = initial_mapping_df.groupby('Source Entity')['Source Entity Description'].first()
    attribute_descriptions = initial_mapping_df.groupby('Source Attribute')['Source Attribute Description'].first()
    key_descriptions = initial_mapping_df.groupby(['Source Entity', 'Keys in Source Entity'])['Key Description'].first()

    glossary = "Source Entity Glossary:\n"
    glossary += "\n".join(f"{entity} - {desc}" for entity, desc in entity_descriptions.items())
    glossary += "\n\nSource Attribute Glossary:\n"
    glossary += "\n".join(f"{attribute} - {desc}" for attribute, desc in attribute_descriptions.items())
    glossary += "\n\nSource Key Constraints:\n"
    glossary += "\n".join(f"{entity}.{key} - {desc}" for (entity, key), desc in key_descriptions.items())
    glossary += f"\n\nTarget Entity Description:\n{initial_mapping_df['Target Entity'].values[0]} - {initial_mapping_df['Target Entity Description'].values[0]}"
    glossary += f"\n\nTarget Key Constraint:\n{pk}"
    glossary += "\n\nTarget Column Glossary:\n"
    glossary += "\n".join(f"{attribute} - {desc}" for attribute, desc in zip(initial_mapping_df['Target Attribute'].unique(), initial_mapping_df['Target Attribute Description'].unique()))

    mapping = "Mapping:\n|Source Entity|Source Attribute|Key Constraint|Target Entity|Target Attribute|\n|-----|-----|-----|-----|\n"
    mapping += "\n".join(f"|{row['Source Entity']}|{row['Source Attribute']}|{row['Target Entity']}|{row['Target Attribute']}|" for _, row in initial_mapping_df.iterrows())

    context = f"{glossary}\n\n{mapping}"

    included_mappings = mapping_chain.invoke({"context": context})
    included_entities = set()
    values_to_match = [
        included_mapping.split('|')[1:5] 
        for included_mapping in included_mappings.split('\n') 
        if '|' in included_mapping and included_mapping.startswith('|')
        if not included_entities.add(included_mapping.split('|')[1])
    ]
    mask_to_keep = pd.Series(False, index=initial_mapping_df.index)
    for values in values_to_match:
        if len(values) == 4:
            current_mask = (initial_mapping_df[columns_to_check] == values).all(axis=1)
            mask_to_keep = mask_to_keep | current_mask
    mapping_df = initial_mapping_df[mask_to_keep]

    included_mappings += "\n\nKey Constraints:\n" + "\n".join(f"{entity}.{key} - {desc}" for (entity, key), desc in key_descriptions.items() if entity in included_entities)
    query = sql_chain.invoke({"context": included_mappings})

    # print(f"Context:\n{context}\n\nMapping Rules:\n{included_mappings}\n\nQuery:\n{query}")

    query_start_index = query.find('```sql')
    query = query.replace('```sql', '')
    query_end_index = query.find('```')
    if query_end_index != -1 and not mapping_df.empty:
        mapping_df.iloc[0, mapping_df.columns.get_loc('Mapping Query')] = query[query_start_index:query_end_index].strip('```')
        return mapping_df
    
    return None

def create_mappings(file):
    c360_df = pd.read_excel(file)
    c360_df = c360_df.fillna('')
    c360_df['Key'] = c360_df['Key'].str.lower()
    unique_entities = c360_df['Entity_Name'].unique()
    entities = [c360_df[c360_df['Entity_Name'] == table] for table in unique_entities]
    c360_df_pk = c360_df[c360_df['Key'] == 'pk']
    c360_df_pk['Key_Description'] = c360_df_pk.apply(lambda row: f"{row['Attribute_Name']} - {row['Attribute_Description']}", axis=1)
    pk_dict = dict(zip(c360_df_pk['Entity_Name'], c360_df_pk['Key_Description']))
    all_rows = []
    progress = 0
    progress_bar = st.progress(progress, f"Retrieving Entities and Attributes from Source System...")

    for entity in entities:
        entity_name = entity.loc[entity['Key'] == "pk", "Entity_Name"].values[0]
        entity_description = entity.loc[entity['Key'] == "pk", "Entity_Description"].values[0]
        pk_name = entity.loc[entity['Key'] == "pk", "Attribute_Name"].values[0].replace('_no', ' number').replace('_id', ' identifier').replace('_', ' ')
        pk_description = entity.loc[entity['Key'] == "pk", "Attribute_Description"].values[0]
        num_attributes = entity.shape[0]
        total_results = 150
        total_results_per_attribute = int(total_results / num_attributes)
        source_entities = search_glossary_entity(entity_description)

        with ThreadPoolExecutor(max_workers=20) as executor:
            futures = []
            for _, attribute in entity.iterrows():
                if attribute['Attribute_Name'] != pk_name:
                    future = executor.submit(
                        lambda attr: (
                            attr, search_glossary_attribute(
                                attr['Attribute_Description'],
                                source_entities,
                                total_results_per_attribute
                            )
                        ), attribute
                    )
                    futures.append(future)
        
        with ThreadPoolExecutor(max_workers=20) as executor:
            processing_futures = [executor.submit(process_future_result, future, pk_name, pk_description) for future in futures]
            for processing_future in as_completed(processing_futures):
                all_rows.extend(processing_future.result())
        
        progress_bar.progress(progress + int(100/len(entities)), text=f"Retrieved Entities and Attributes from Source System for {entity} ...")
    progress_bar.empty()

    all_initial_mapping_df = pd.DataFrame(all_rows, columns=[
        'Source System',
        'Source Entity', 
        'Source Entity Description', 
        'Source Attribute', 
        'Source Attribute Description',
        'Keys in Source Entity',
        'Key Description', 
        'Target Entity', 
        'Target Entity Description',
        'Target Attribute',
        'Target Attribute Description',
        'Target Attribute Data Type',
        'Mapping Query'
    ])
    # all_initial_mapping_df.to_excel(f"../output/mapping/{file.name} mapping.xlsx", sheet_name="Initial", header=True, index=False)

    target_entities = all_initial_mapping_df['Target Entity'].unique()
    initial_mapping_df_list = [all_initial_mapping_df[all_initial_mapping_df['Target Entity'] == target_entity] for target_entity in target_entities]

    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_df = {executor.submit(process_mapping_df, df, pk_dict[df['Target Entity'].values[0]]): df for df in initial_mapping_df_list}
        mapping_df_list = []
        for future in as_completed(future_to_df):
            mapping_df_list.append(future.result())

    return pd.concat(mapping_df_list, ignore_index=True)

def create_entity_label(entity, attributes):
    return f"{entity}\\n" + "\\l".join(attributes) + "\\l"

def visualise_mapping(df, name):
    dot = Digraph(node_attr={'shape': 'box', 'style': 'filled'})
    source_entities = df.groupby('Source Entity')['Source Attribute'].apply(list).to_dict()
    target_entities = df.groupby('Target Entity')['Target Attribute'].apply(list).to_dict()

    source_color = '#D0E8FF'  # Light orange
    target_color = '#B3DE69'  # Light green

    for entity, attributes in source_entities.items():
        dot.node(entity, create_entity_label(entity, attributes), fillcolor=source_color)

    for entity, attributes in target_entities.items():
        dot.node(entity, create_entity_label(entity, attributes), fillcolor=target_color)

    for _, row in df.iterrows():
        dot.edge(row['Source Entity'], row['Target Entity'], color='gray')

    dot.render(f"./output/mapping/{name} mapping", format='png', cleanup=True)
    return f"./output/mapping/{name} mapping.png"

aggregation_prompt = PromptTemplate.from_template(
""" 
You are a data scientist with expertise in the life insurance domain.

I will provide you with a set of base entities and attributes for each base entity accompanied by the data type and key constraints associated 
with each attribute. Your task is to write the SQL query to aggregate the relevant base entities and attributes to each of the given
target entity from {view} and create a table for the target entity. You will analyse the name of the aggregate view attribute, the names, datatypes and key constraints in the 
set of base entities and attributes available to you to determine the business logic that can be applied to aggregate the base entities and
attributes into the target entity. 

Following are the base entities and attributes:
{base}

Following are the target attributes and optionally, entities:
{derived}

You will respond in the following format:
| Aggregate View Entity | SQL Query |

Respond only with the table according to the output format. Do not include introductions, explanations, conclusions, SQL code label and backticks.
"""
)

aggregation_chain = runnable | aggregation_prompt | llm | output_parser

def create_aggregate_view(file, name, aggregate_attributes):
    base_attributes_df = pd.read_excel(file)
    base_attributes = ""
    tables = base_attributes_df['Entity_Name'].unique()
    for table in tables:
        columns = base_attributes_df[base_attributes_df['Entity_Name'] == table].drop('Entity_Name', axis=1).to_string(index=False, header=False)
        base_attributes += f"Table: {table}\n{columns}\n\n"
    
    query = aggregation_chain.invoke({"view": name, "base": base_attributes, "derived": aggregate_attributes})

    try:
        query_table = query.splitlines()
        aggregation_query_df = pd.DataFrame([row.split('|')[1:-1] for row in query_table[2:]], columns=query_table[0].split('|')[1:-1])
        return aggregation_query_df 
    except:
        return None

# Streamlit Interface
# if 'is_indexed' not in st.session_state:
#     st.session_state['is_indexed'] = False
if 'metadata_list' not in st.session_state:
    st.session_state['metadata_list'] = None
if 'table_data_list' not in st.session_state:
    st.session_state['table_data_list'] = None
if 'base_mapping' not in st.session_state:
    st.session_state['base_mapping'] = None
if 'base_mapping_image' not in st.session_state:
    st.session_state['base_mapping_image'] = None
if 'aggregate_mapping' not in st.session_state:
    st.session_state['aggregate_mapping'] = None
# if 'aggregate_attributes' not in st.session_state:
#     st.session_state['aggregate_attributes'] = None
if 'aggregate_name' not in st.session_state:
    st.session_state['aggregate_name'] = None


st.set_page_config(
    layout="wide",
    page_title="Automatic Source to Target Mapping",
    page_icon="↔️",
)

st.subheader("Automatically Map Attributes in Source System to Attributes in a Data Model using AI ↔️")
st.markdown(
"""
This tool expedites the Mapping of attributes in source systems to attributes in any data model. You may upload an excel workbook
containing a data dictionary (Entity Name, Entity Description, Attribute Name, Attribute Description, Key Constraints, Data Type)
of the target data model. You may select multiple source systems from a list of popular transactional databases in the BFSI sector
or upload a data dictionary of a source system. 

The outcome is an Excel Workbook containing a mapping table and the SQL query to perform the mapping. The workbook can be found in the 
**output/mapping** directory.

----
""")

with st.sidebar:
    st.write("### Source System")
    source_file = st.file_uploader("Upload Dictionary of Source System")
    if source_file and st.session_state['base_mapping'] is None:
        with st.spinner("Indexing the Dictionary (This may take a while)..."):
            new_metadata_list, new_table_data_list = create_index(source_file)
            if new_metadata_list and new_table_data_list:
                metadata_list.extend(new_metadata_list)
                st.session_state['metadata_list'] = metadata_list
                table_data_list.extend(new_table_data_list)
                st.session_state['table_data_list'] = table_data_list
                st.session_state['is_indexed'] = True
                st.success(f"Index created for {source_file.name}")

st.write("#### Base Model")
data_model_file = st.file_uploader("Upload Dictionary of Base Model", "xlsx")

st.write("#### Aggregate Model")
name = st.text_input("Please enter the name of the aggregated view")
aggregate_attributes = st.text_area("Please enter the derived attributes", height=200)

if metadata_list != [] and data_model_file is not None:
    if st.button("Map Source System to Base Attributes"):
        base_mapping_df = create_mappings(data_model_file)
        if base_mapping_df is not None:
            st.session_state['base_mapping'] = base_mapping_df
        else:
            st.error("Something went wrong! Please try again.")

if data_model_file and name.strip() != "" and aggregate_attributes.strip() != "":
    if st.button("Map Base Attributes to Aggregate Attributes"):
        st.session_state['aggregate_name'] = name

        with st.spinner("Mapping Base to Aggregate...."):
            aggregate_mapping_df = create_aggregate_view(data_model_file, name, aggregate_attributes)
            if aggregate_mapping_df is not None:
                st.session_state['aggregate_mapping'] = aggregate_mapping_df
            else:
                st.error("Something went wrong! Please try again.")

if st.session_state['base_mapping'] is not None:
    st.subheader("Mapping and Queries", divider='violet')
    st.dataframe(st.session_state['base_mapping'])
    # entities = [st.session_state['base_mapping'][st.session_state['base_mapping']['Target Entity'] == table] for table in st.session_state['base_mapping']['Target Entity'].unique()]
    # for entity in entities:
    #     with st.expander(entity.at[1, 'Target Entity']):
    #         st.write(f"**Target Entity Description**: {entity.at[1, 'Target Entity Description']}")
    #         st.dataframe(entity[['Source Entity', 'Source Entity Description', 'Source Attribute', 'Source Attribute Description', 'Target Attribute', 'Target Attribute Description']])
    #         st.code(st.session_state['base_mapping'])
    # image_path = visualise_mapping(st.session_state['base_mapping'], data_model_file.name.replace(".xlsx", ""))
    # st.session_state['base_mapping_image'] = image_path
    # if st.session_state['base_mapping_image'] is not None:
    #     st.image(st.session_state['base_mapping_image'], caption="Mapping Diagram")

    if st.button("Save Mappings"):
        st.session_state['base_mapping'].to_excel(f"./output/mapping/{data_model_file.name.replace('.xlsx', '')} mapping.xlsx", sheet_name='Mapping', header=True, index=False)
        # if st.session_state['base_mapping_image'] is not None:
        #     wb = load_workbook(f"./output/mapping/{data_model_file.name.replace(".xlsx", "")} mapping.xlsx")
        #     ws = wb.create_sheet('Diagram')
        #     image = Image(image_path)
        #     ws.add_image(image, 'A1')
        st.success("Mappings saved to **output/mapping**")

if st.session_state['aggregate_mapping'] is not None:
    st.subheader("Mapping and Queries", divider='violet')
    for _, row in st.session_state['aggregate_mapping'].iterrows():
        st.write(f"**{row['Aggregate View Attribute']}**")
        st.code(row['SQL Query'])
    # st.dataframe(st.session_state['aggregate_mapping'])
    if st.button("Save Mappings"):
        st.session_state['aggregate_mapping'].to_excel(f"./output/mapping/{st.session_state['aggregate_name']} mapping.xlsx", sheet_name='Mapping', header=True, index=False)
        st.success("Mappings saved to **output/mapping**")    



