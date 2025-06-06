import streamlit as st
import pandas as pd
import json
from tabulate import tabulate
import os
import regex as re
from typing import List, Optional, Any, Dict, Iterable, Sequence, Union
from typing_extensions import Literal
from deprecated import deprecated

import sqlalchemy
from sqlalchemy import create_engine, inspect, MetaData, Table, Column, Identity, text, select, URL
from sqlalchemy.engine import Engine, Connection, Result
from sqlalchemy.exc import SQLAlchemyError, ProgrammingError
from sqlalchemy.schema import CreateTable
from sqlalchemy.sql.elements import Executable
from sqlalchemy.sql.type_api import TypeEngine
from sqlalchemy.types import NullType
from sqlalchemy.util.typing import Literal as DepLiteral
from sqlalchemy.engine.url import make_url
from sqlalchemy_utils import database_exists
from urllib.parse import quote_plus

from langchain_community.utilities import SQLDatabase as _SQLDatabase
from langchain_core.callbacks import CallbackManagerForToolRun
from langchain.agents import create_sql_agent
from langchain_openai import AzureChatOpenAI
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser

from sdv.utils import drop_unknown_references
from sdv.multi_table import HMASynthesizer
from sdv.metadata import Metadata
from table_evaluator import TableEvaluator
from sdmetrics.reports.multi_table import QualityReport, DiagnosticReport

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


llm = AzureChatOpenAI(
    azure_deployment="gpt-4o",
    openai_api_version="2024-02-15-preview",
    api_key="901cce4a4b4045b39727bf1ce0e36219",
    azure_endpoint="https://aifordata-azureopenai.openai.azure.com/",
    temperature=0,
    max_tokens=4096
)

runnable = RunnablePassthrough()
output_parser = StrOutputParser()


pd.set_option("styler.render.max_elements", 5000000)



def handle_upload(file, tables=[]):
    hci_data = {}
    if file.name.endswith('.csv'):
        hci_data = {file.name.replace(".csv", ""): pd.read_csv(file)}
        return hci_data
    elif file.name.endswith('.xlsx') or file.name.endswith('.xls'):
        if tables != []:
            for i, sheet in enumerate(tables):
                data = pd.read_excel(file, sheet_name=sheet)
                hci_data[sheet] = data
                if i == 5:
                    break
            return hci_data
        else:
            workbook = load_workbook(file)
            if workbook:
                for i, sheet in enumerate(workbook.sheetnames):
                    data = pd.read_excel(file, sheet_name=sheet)
                    hci_data[sheet] = data
                    if i == 5:
                        break
                return hci_data
    else:
        st.warning("Selected file is not in a supported format.")

    return None

def generate_metadata(hci_data):
    metadata = Metadata.detect_from_dataframes(hci_data)
    metadata_dict = metadata.to_dict()
    for table, table_metadata in metadata_dict['tables'].items():
        for field, field_metadata in table_metadata['columns'].items():
            if 'pii' in field_metadata and field_metadata['pii'] == True:
                del field_metadata['pii']
    metadata = Metadata.load_from_dict(metadata_dict)
    return metadata

def model_hma_synthesizer(hci_data, metadata, number_of_rows):
    synthesizer = HMASynthesizer(metadata, locales=[])
    synthesizer.fit(hci_data)
    row_count = min(df.shape[0] for df in hci_data.values())
    scale = float(number_of_rows/row_count)
    synthetic_data = synthesizer.sample(scale=scale)

    # for relationship in metadata['relationships']:
    #     parent_table_name = relationship['parent_table']
    #     child_table_name = relationship['child_table']
    #     foreign_key_column = relationship['child_table_foreign_key']
    #     primary_key_column = relationship['parent_table_primary_key']

    return synthetic_data

def backfill_synthetic_data(synthetic_data, specification, description):
    synthetic_data = synthetic_data.map(lambda x: x.replace("sdv-id-", "") if isinstance(x, str) and 'sdv-id' in x else x)
    profile = {
        'number_of_rows': synthetic_data.shape[0],
        'column_names_and_data_types': synthetic_data.dtypes.to_dict(),
        'number_of_missing_values': synthetic_data.isnull().sum().to_dict(),
        'unique_categorical_values': {col: [synthetic_data[col].unique().tolist()] for col in synthetic_data.select_dtypes(include=['category']).columns.tolist()},
        'date_ranges': {col: f"{synthetic_data[col].min().to_pydatetime()} to {synthetic_data[col].max().to_pydatetime()}" for col in synthetic_data.select_dtypes(include=['datetime']).columns.tolist()},
        'numerical_ranges': {col: f"{synthetic_data[col].dropna().min()} to {synthetic_data[col].dropna().max()}" for col in synthetic_data.select_dtypes(include=['number']).columns.tolist()},
        'statistical_properties_of_numerical_fields': tabulate(synthetic_data.describe(include=['number']), headers='keys', tablefmt='simple_grid')
    }
    print(description)
    print(profile)

    null_before = synthetic_data.isnull().sum().sum()

    script_prompt = PromptTemplate.from_template(
    """
    You are an expert in data analysis and enhancing synthetic data. Your task is to create a python script to logically backfill null values
    in synthetic test data contained within a dataframe by understanding various information about the data and a JIRA story or similar specification
    document of test scnearios which will be provided to you as context. Optionally you will also be provided with a glossary of columns in case
    of ambiguous and codified column names.

    You will adhere to the following steps in order to guide your creation of the script:
    Thoroughly parse the specification document, information about the data and column glossary (if provided).
    You will first determine the columns in the dataframe that are relevant to the use case.
    You will analyse the names of the columns and their descriptions (if provided), analyse the constraints and formats specified in the test case
    specification, information about the numerical and date time ranges, valid categorical values and statistical properties of numerical values,
    then augment your analysis by drawing context from the test scenario specifications.
    You will thoroughly reason about logical relationships among the fields and the data under them.
    You will not backfill fields that contain no data under them (same number of nulls as number of rows in the dataframe).
    You will distinguish nulls which need to be backfilled and nulls which should not be backfilled based on your understanding of logical possibility
    of a field being blank (for eg s certain flag value may negate the possibility of values being available for consequent fields).
    You will use your reasoning to mine patterns, logic and rules that may not be explicitly specified in the test specifications, arrive at a detailed
    reasoning and approach towards generating the python script.
    You may also define a dictionary of categorical values to better guide the synthetic data generation process in case of interdependent fields (for eg,
    category name and category code fields may be interdependent).
    You will generate a highly performant, fool-proof and fault tolerant python script that will logically backfill null values in the relevant columns of the
    dataframe with synthetic data using the faker library within the constraints of the values of categorical fields, numerical ranges, date ranges and all test
    specifications.
    The script should not affect the data under fields that are not relevant to the test scenarios.
    You will not create a new dataframe. You will not initialise a dataframe with sample values. I have already created the dataframe named 'df'. 
    You will provide me with the python script that I can execute to perform backfilling on the 'df' dataframe which already exists, and accomplish the task.
    You will not include any statements for printing the df.
    Finally, perform any necessary data type conversions to restore the integrity of the dataframe, accurate specifications of formats and coercion
    in case of error.

    Following is the information about the dataset:
    {profile}

    Optionally a description of the columns:
    {description}

    Following is the test case specification:
    {specification}

    Respond in the following format:
    Relevant Columns:
    [list of relevant columns]

    Python Script:
    [python script]
    """)

    fix_code_prompt = PromptTemplate.from_template(
    """
    You are an incredible software engineer with expertise in fixing and debugging python programs. I tried executing the following code:
    {code}

    And received the following error:
    {error}

    Understand the goal of the python program, the cause of error and fix the code without deviating from the goal.
    Respond with the complete fixed code.
    """
    )

    script_chain = runnable | script_prompt | llm | output_parser
    fix_code_chain = runnable | fix_code_prompt | llm | output_parser

    script = script_chain.invoke({
        "profile": profile,
        "specification": specification,
        "description": description
    })
    # print(script)
    start_index = script.find("```python")
    script = script.replace("```python", "")
    end_index = script.find("```")
    code = script[start_index: end_index]
    df = synthetic_data
    print(code)

    i = 0
    error = True
    while i <= 5 and error:
        try:
            local_vars = locals()
            exec(code, local_vars, None)
            df = local_vars['df']
            null_after = df.isnull().sum().sum()
            return null_before, null_after, df
        except Exception as e:
            print(e)
            null_after = df.isnull().sum().sum()
            i += 1
            error = True
            code = fix_code_chain.invoke({"code": code, "error": e})
            start_index = code.find("```python")
            code = code.replace("```python", "")
            end_index = code.find("```")
            code = code[start_index: end_index]
            print(code)
            continue

    return null_before, null_after, synthetic_data

@st.cache_data
def table_evaluator(hci_data, synthetic_data):
    table_evaluator = TableEvaluator(hci_data, synthetic_data)
    return table_evaluator

def get_connection_string(config):
    drivers = {
        "Microsoft SQL Server": "mssql+pyodbc",
        "MySQL": "mysql+pymysql",
        "OracleDB": "oracle+cx_oracle",
        "SQLite": "sqlite",
        "PostgreSQL": "postgresql",
    }

    if 'password' in config:
        config['password'] = quote_plus(config['password'])

    drivername = drivers.get(config['rdbms'])
    if not drivername:
        return f"Unsupported RDBMS: {config['rdbms']}"

    try:
        if config['rdbms'] == "Microsoft SQL Server":
            if config['authentication'] == "Windows Authentication":
                conn_str = f"{drivername}://{config['host']}/{config['database']}?driver=ODBC+Driver+17+for+SQL+Server&Trusted_Connection=yes"
            else:
                conn_str = f"{drivername}://{config['username']}:{config['password']}@{config['host']}/{config['database']}?driver=ODBC+Driver+17+for+SQL+Server&charset=utf16"
        elif config['rdbms'] == "OracleDB":
            conn_str = f"{drivername}://{config['username']}:{config['password']}@{config['host']}:{config['port']}/?service_name={config['service']}&encoding=UTF-8&nencoding=UTF-8"
        elif config['rdbms'] == "SQLite":
            config['database'] = config['database'].strip('"').replace('\\','/')
            if database_exists(f"{drivername}:///{config['database']}"):
                conn_str = f"{drivername}:///{config['database']}"
        else:
            conn_str = URL.create(
                drivername=drivername,
                username=config['username'],
                password=config['password'],
                host=config['host'],
                port=config['port'],
                database=config['database']
            )
        return conn_str
    except Exception as e:
        return None

def get_schemas(conn_str):
    try:
        if conn_str.startswith('sqlite:'):
            return ['main']
        
        engine = create_engine(conn_str)
        inspector = inspect(engine)
        schemas = inspector.get_schema_names()

        system_schemas = ['INFORMATION_SCHEMA', 'sys',
                          'information_schema', 'mysql', 'performance_schema',
                          'SYS', 'SYSTEM', 'OUTLN', 'CTXSYS', 'MDSYS']
        
        schemas = [schema for schema in schemas if 
                   schema not in system_schemas 
                   and not schema.startswith('db_') 
                   and not schema.startswith('pg_')]
        
        return schemas
    except Exception as e:
        return f"Error fetching schemas: {e}"  

def get_business_glossary(file_path):
    dd_df = pd.read_excel(file_path, engine='openpyxl')
    dd_df['Schema Table'] = dd_df['Schema'] + '.' + dd_df['Table'] 
    table_description = dd_df.groupby('Schema Table')['Table Description'].first().to_dict()
    column_description = dd_df.groupby('Schema Table').apply(
        lambda group: '\n'.join(f"{row['Column']} - {row['Definition']}" for _, row in group.iterrows())
    ).to_dict()
    return table_description, column_description

def _format_index(index: sqlalchemy.engine.interfaces.ReflectedIndex) -> str:
    return (
        f'Name: {index["name"]}, Unique: {index["unique"]},'
        f' Columns: {str(index["column_names"])}'
    )

def truncate_word(content: Any, *, length: int, suffix: str = "...") -> str:
    """
    Truncate a string to a certain number of words, based on the max string
    length.
    """

    if not isinstance(content, str) or length <= 0:
        return content

    if len(content) <= length:
        return content

    return content[: length - len(suffix)].rsplit(" ", 1)[0] + suffix

class SQLDatabase(_SQLDatabase):
    """SQLAlchemy wrapper around a database with modified get_data and get_metadata."""
    def __init__(
        self,
        engine: Engine,
        schemas: List[str],
        metadata: Optional[MetaData] = None,
        ignore_tables: Optional[List[str]] = None,
        include_tables: Optional[List[str]] = None,
        sample_rows_in_table_info: int = 2,
        indexes_in_table_info: bool = False,
        custom_table_info: Optional[dict] = None,
        table_description: Optional[dict] = None,
        column_description: Optional[dict] = None,
        view_support: bool = False,
        max_string_length: int = 300,
        lazy_table_reflection: bool = False,
    ):
        """Create engine from database URI."""
        self._engine = engine
        self._schemas = schemas
        if include_tables and ignore_tables:
            raise ValueError("Cannot specify both include_tables and ignore_tables")

        self._inspector = inspect(self._engine)

        self._all_tables_per_schema = {}
        for schema in self._schemas:
            self._all_tables_per_schema[schema] = set(
                self._inspector.get_table_names(schema=schema)
                + (self._inspector.get_view_names(schema=schema) if view_support else [])
            )
        if table_description:
            self._all_tables = set(f"{k}.{name} - {table_description.get(f"{k}.{name}", "")}" for k, names in self._all_tables_per_schema.items() for name in names)
        else:
            self._all_tables = set(f"{k}.{name}" for k, names in self._all_tables_per_schema.items() for name in names)
        

        self._include_tables = set(include_tables) if include_tables else set()
        if self._include_tables:
            missing_tables = self._include_tables - self._all_tables
            if missing_tables:
                raise ValueError(
                    f"include_tables {missing_tables} not found in database"
                )
        self._ignore_tables = set(ignore_tables) if ignore_tables else set()
        if self._ignore_tables:
            missing_tables = self._ignore_tables - self._all_tables
            if missing_tables:
                raise ValueError(
                    f"ignore_tables {missing_tables} not found in database"
                )
        usable_tables = self.get_usable_table_names()
        self._usable_tables = set(usable_tables) if usable_tables else self._all_tables

        if not isinstance(sample_rows_in_table_info, int):
            raise TypeError("sample_rows_in_table_info must be an integer")

        self._sample_rows_in_table_info = sample_rows_in_table_info
        self._indexes_in_table_info = indexes_in_table_info

        self._custom_table_info = custom_table_info
        if self._custom_table_info:
            if not isinstance(self._custom_table_info, dict):
                raise TypeError(
                    "table_info must be a dictionary with table names as keys and the "
                    "desired table info as values"
                )
            # only keep the tables that are also present in the database
            intersection = set(self._custom_table_info).intersection(self._all_tables)
            self._custom_table_info = dict(
                (table, self._custom_table_info[table])
                for table in self._custom_table_info
                if table in intersection
            )

        # Initialising column descriptions dictionary
        self._column_description = column_description or {}

        self._max_string_length = max_string_length
        self._view_support = view_support

        self._metadata = metadata or MetaData()
        if not lazy_table_reflection:
            for schema in self._schemas:
                self._metadata.reflect(
                    views=view_support,
                    bind=self._engine,
                    schema=schema,
                )
        # Add id to tables metadata
        for t in self._metadata.sorted_tables:
            t.id = f"{t.schema}.{t.name}"

        self._table_data = {}
        self._table_metadata = {}

    @classmethod
    def from_uri(
        cls, database_uri: str, engine_args: Optional[dict] = None, schemas: Optional[List[str]] = None, **kwargs: Any
    ):
        """Construct a SQLAlchemy engine from URI."""
        _engine_args = engine_args or {}
        return cls(create_engine(database_uri, **_engine_args), schemas=schemas, **kwargs)

    @property
    def dialect(self) -> str:
        """Return string representation of dialect to use."""
        return self._engine.dialect.name

    def get_usable_table_names(self) -> Iterable[str]:
        """Get names of tables available."""
        if self._include_tables:
            return sorted(self._include_tables)
        return sorted(self._all_tables - self._ignore_tables)

    # @deprecated("0.0.1", alternative="get_usable_table_names", removal="0.3.0")
    # def get_table_names(self) -> Iterable[str]:
    #     """Get names of tables available."""
    #     return self.get_usable_table_names()

    @property
    def table_info(self) -> str:
        """Information about all tables in the database."""
        return self.get_table_info()

    def get_table_info(self, table_names: Optional[List[str]] = None) -> str:
        """Get information about specified tables.

        Follows best practices as specified in: Rajkumar et al, 2022
        (https://arxiv.org/abs/2204.00498)

        If `sample_rows_in_table_info`, the specified number of sample rows will be
        appended to each table description. This can increase performance as
        demonstrated in the paper.
        """
        all_table_names = self.get_usable_table_names()
        all_table_names = [table_name.split(" - ")[0].strip("'") if " - " in table_name else table_name.strip("'") for table_name in all_table_names]

        if table_names is not None:
            table_names = [table_name.split(" - ")[0].strip("'") if " - " in table_name else table_name.strip("'") for table_name in table_names]
            missing_tables = set(table_names).difference(all_table_names)
            if missing_tables:
                raise ValueError(f"table_names {missing_tables} not found in database")
            all_table_names = table_names

        metadata_table_names = [tbl.name for tbl in self._metadata.sorted_tables]
        to_reflect = set(all_table_names) - set(metadata_table_names)
        if to_reflect:
            for schema in self._schemas:
                self._metadata.reflect(
                    views=self._view_support,
                    bind=self._engine,
                    schema=schema,
                )

        meta_tables = [
            tbl
            for tbl in self._metadata.sorted_tables
            if tbl.id in set(all_table_names)
            and not (self.dialect == "sqlite" and tbl.name.startswith("sqlite_"))
        ]

        tables = []
        for table in meta_tables:
            if self._custom_table_info and table.name in self._custom_table_info:
                tables.append(self._custom_table_info[table.name])
                continue

            # Ignore JSON datatyped columns
            for k, v in table.columns.items():
                if type(v.type) is NullType:
                    table._columns.remove(v)

            # add create table command
            create_table = str(CreateTable(table).compile(self._engine))
            table_info = f"{create_table.rstrip()}"
            column_desc = self._column_description.get(table.id)
            if column_desc:
                table_info += "\n\n/* Column Descriptions:\n"
                table_info += column_desc
                table_info += "\n*/"     
            has_extra_info = (
                self._indexes_in_table_info or self._sample_rows_in_table_info
            )
            if has_extra_info:
                table_info += "\n\n/*"
            if self._indexes_in_table_info:
                table_info += f"\n{self._get_table_indexes(table)}\n"
            if self._sample_rows_in_table_info:
                table_info += f"\n{self._get_sample_rows(table)}\n"
            if has_extra_info:
                table_info += "*/"

            tables.append(table_info)
            
        tables.sort()
        final_str = "\n\n".join(tables)
        return final_str

    def _get_table_indexes(self, table: Table) -> str:
        indexes = self._inspector.get_indexes(table.name)
        indexes_formatted = "\n".join(map(_format_index, indexes))
        return f"Table Indexes:\n{indexes_formatted}"

    def _get_sample_rows(self, table: Table) -> str:
        # build the select command
        command = select(table).limit(self._sample_rows_in_table_info)

        # save the columns in string format
        columns_str = "\t".join([col.name for col in table.columns])

        try:
            # get the sample rows
            with self._engine.connect() as connection:
                sample_rows_result = connection.execute(command)  # type: ignore
                # shorten values in the sample rows
                sample_rows = list(
                    map(lambda ls: [str(i)[:100] for i in ls], sample_rows_result)
                )

            # save the sample rows in string format
            sample_rows_str = "\n".join(["\t".join(row) for row in sample_rows])

        # in some dialects when there are no rows in the table a
        # 'ProgrammingError' is returned
        except ProgrammingError:
            sample_rows_str = ""

        return (
            f"{self._sample_rows_in_table_info} rows from {table.name} table:\n"
            f"{columns_str}\n"
            f"{sample_rows_str}"
        )

    def _execute(
            self,
            command: Union[str, Executable],
            fetch: Literal["all", "one", "cursor"] = "all",
            *,
            parameters: Optional[Dict[str, Any]] = None,
            execution_options: Optional[Dict[str, Any]] = None,
        ) -> Union[Sequence[Dict[str, Any]], Result]:
            """
            Executes SQL command through underlying engine.

            If the statement returns no rows, an empty list is returned.
            """
            parameters = parameters or {}
            execution_options = execution_options or {}
            with self._engine.begin() as connection:  # type: Connection  # type: ignore[name-defined]
                # if self._schema is not None:
                #     if self.dialect == "snowflake":
                #         connection.exec_driver_sql(
                #             "ALTER SESSION SET search_path = %s",
                #             (self._schema,),
                #             execution_options=execution_options,
                #         )
                #     elif self.dialect == "bigquery":
                #         connection.exec_driver_sql(
                #             "SET @@dataset_id=?",
                #             (self._schema,),
                #             execution_options=execution_options,
                #         )
                #     elif self.dialect == "mssql":
                #         pass
                #     elif self.dialect == "trino":
                #         connection.exec_driver_sql(
                #             "USE ?",
                #             (self._schema,),
                #             execution_options=execution_options,
                #         )
                #     elif self.dialect == "duckdb":
                #         # Unclear which parameterized argument syntax duckdb supports.
                #         # The docs for the duckdb client say they support multiple,
                #         # but `duckdb_engine` seemed to struggle with all of them:
                #         # https://github.com/Mause/duckdb_engine/issues/796
                #         connection.exec_driver_sql(
                #             f"SET search_path TO {self._schema}",
                #             execution_options=execution_options,
                #         )
                #     elif self.dialect == "oracle":
                #         connection.exec_driver_sql(
                #             f"ALTER SESSION SET CURRENT_SCHEMA = {self._schema}",
                #             execution_options=execution_options,
                #         )
                #     elif self.dialect == "sqlany":
                #         # If anybody using Sybase SQL anywhere database then it should not
                #         # go to else condition. It should be same as mssql.
                #         pass
                #     elif self.dialect == "postgresql":  # postgresql
                #         connection.exec_driver_sql(
                #             "SET search_path TO %s",
                #             (self._schema,),
                #             execution_options=execution_options,
                #         )

                if isinstance(command, str):
                    command = text(command)
                elif isinstance(command, Executable):
                    pass
                else:
                    raise TypeError(f"Query expression has unknown type: {type(command)}")
                cursor = connection.execute(
                    command,
                    parameters,
                    execution_options=execution_options,
                )

                if cursor.returns_rows:
                    if fetch == "all":
                        result = [x._asdict() for x in cursor.fetchall()]
                    elif fetch == "one":
                        first_result = cursor.fetchone()
                        result = [] if first_result is None else [first_result._asdict()]
                    elif fetch == "cursor":
                        return cursor
                    else:
                        raise ValueError(
                            "Fetch parameter must be either 'one', 'all', or 'cursor'"
                        )
                    return result
            return []

    def run(
            self,
            command: Union[str, Executable],
            fetch: Literal["all", "one", "cursor"] = "all",
            include_columns: bool = False,
            parameters: Optional[Dict[str, Any]] = None,
            execution_options: Optional[Dict[str, Any]] = None
        ) -> Union[str, Sequence[Dict[str, Any]], Result[Any]]:
            """Executes SQL command to retrieve relevant data from each relevant table and returns a string indicating whether the data was successfully retrieved. 
            """            
            result = self._execute(
                command, fetch, parameters=parameters, execution_options=execution_options
            )
            if len(result) > 0:
                return f"This query will fetch {len(result)} rows of data."
            else:
                return "This query will not retrieve data. Please try again."

    def get_table_info_no_throw(self, table_names: Optional[List[str]] = None) -> str:
        """Get information about specified tables.

        Follows best practices as specified in: Rajkumar et al, 2022
        (https://arxiv.org/abs/2204.00498)

        If `sample_rows_in_table_info`, the specified number of sample rows will be
        appended to each table description. This can increase performance as
        demonstrated in the paper.
        """
        try:
            return self.get_table_info(table_names)
        except ValueError as e:
            """Format the error message"""
            return f"Error: {e}"

    def run_no_throw(
        self,
        command: str,
        fetch: Literal["all", "one"] = "all",
        include_columns: bool = False,
        *,
        parameters: Optional[Dict[str, Any]] = None,
        execution_options: Optional[Dict[str, Any]] = None,
    ) -> Union[str, Sequence[Dict[str, Any]], Result[Any]]:
        """Execute a SQL command and return a string representing the results.

        If the statement returns rows, a string of the results is returned.
        If the statement returns no rows, an empty string is returned.

        If the statement throws an error, the error message is returned.
        """
        try:
            return self.run(
                command,
                fetch,
                parameters=parameters,
                execution_options=execution_options,
                include_columns=include_columns,
            )
        except SQLAlchemyError as e:
            """Format the error message"""
            return f"Error: {e}"

    def get_context(self) -> Dict[str, Any]:
        """Return db context that you may want in agent prompt."""
        table_names = list(self.get_usable_table_names())
        table_info = self.get_table_info_no_throw()
        return {"table_info": table_info, "table_names": ", ".join(table_names)}
 
    def table_data(
            self,
            command: Union[str, Executable],
            fetch: Literal["all", "one", "cursor"] = "all",
            include_columns: bool = False,
            parameters: Optional[Dict[str, Any]] = None,
            execution_options: Optional[Dict[str, Any]] = None
        ):
        self._get_table_data(
                command,
                fetch,
                parameters=parameters,
                execution_options=execution_options,
                include_columns=include_columns
        )
        return self._table_data
    
    def _get_table_data(            
            self,
            command: Union[str, Executable],
            fetch: Literal["all", "one", "cursor"] = "all",
            include_columns: bool = False,
            parameters: Optional[Dict[str, Any]] = None,
            execution_options: Optional[Dict[str, Any]] = None
        ):
        for table_command in command.split("\n\n"):
            result = self._execute(
                table_command, fetch, parameters=parameters, execution_options=execution_options
            )
            if result:
                match = re.search(r'FROM\s+([^\s;]+)', table_command, re.IGNORECASE)
                if match:
                    table_name = match.group(1)

                    for row in result:
                        for key, value in row.items():
                            if isinstance(value, bytes):
                                try:
                                    row[key] = value.decode('utf-8')
                                except UnicodeDecodeError:
                                    row[key] = value.decode('iso-8859-1')

                    self._table_data[table_name] = pd.DataFrame(result)

    @property
    def table_metadata(self):
        return self._get_metadata()
    
    def _get_metadata(self) -> Dict:
        metadata_dict = {
            "METADATA_SPEC_VERSION": "V1",
            "tables": {},
            "relationships": []
        }
        table_names_schemas = self._table_data.keys()

        for table_name_schema in table_names_schemas:
            schema, table_name = table_name_schema.split('.')
            sqlalchemy_table = Table(table_name, self._metadata, schema=schema, autoload_with=self._engine)

            columns_metadata = {}
            primary_key_columns = [col.name for col in sqlalchemy_table.primary_key.columns]
            foreign_key_columns = [fk.parent.name for fk in sqlalchemy_table.foreign_keys]

            for column in sqlalchemy_table.columns:
                if column.name in primary_key_columns or column.name in foreign_key_columns:
                    col_metadata = {"sdtype": "id"}
                else:
                    sdtype = self._map_sqlalchemy_type_to_sdtype(column.type)
                    col_metadata = {"sdtype": sdtype}
                if col_metadata["sdtype"] == "id":
                    regex_format = self._infer_regex_format(column) 
                    if regex_format:
                        col_metadata["regex_format"] = regex_format
                elif sdtype == "datetime":
                    datetime_format = self._infer_datetime_format(column)
                    if datetime_format:
                        col_metadata["datetime_format"] = datetime_format
                columns_metadata[column.name] = col_metadata

            metadata_dict["tables"][table_name_schema] = {
                "primary_key": primary_key_columns[0] if primary_key_columns else None, 
                "columns": columns_metadata,
                "column_relationships": [] 
            }

        for table_name_schema in table_names_schemas:
            schema, table_name = table_name_schema.split('.')
            sqlalchemy_table = Table(table_name, self._metadata, schema=schema, autoload_with=self._engine)
            for fk_constraint in sqlalchemy_table.foreign_key_constraints:
                for fk in fk_constraint.elements:
                    parent_table_name_schema = f"{fk.column.table.schema}.{fk.column.table.name}" if fk.column.table.schema else fk.column.table.name
                    child_table_name_schema = table_name_schema
                    if parent_table_name_schema in table_names_schemas and child_table_name_schema in table_names_schemas:
                        metadata_dict["relationships"].append({
                            "parent_table_name": parent_table_name_schema,
                            "parent_primary_key": fk.column.name,
                            "child_table_name": child_table_name_schema,
                            "child_foreign_key": fk.parent.name
                        })
        return metadata_dict

    def _map_sqlalchemy_type_to_sdtype(self, sqlalchemy_type: TypeEngine) -> str:
        """Maps SQLAlchemy types to SDV sdtypes."""
        type_name = sqlalchemy_type.__visit_name__.lower()
        if "int" in type_name or "numeric" in type_name or "float" in type_name or "decimal" in type_name:
            return "numerical"
        elif "char" in type_name or "text" in type_name or "string" in type_name:
            return "categorical" 
        elif "date" in type_name or "time" in type_name:
            return "datetime"
        elif "bool" in type_name:
            return "boolean"
        else:
            return "categorical" 

    def _infer_regex_format(self, column: Column) -> Optional[str]:
        """Infers regex format for ID columns (basic heuristic)."""
        if column.name.lower().endswith("_id"):
            return r".*" 
        return None

    def _infer_datetime_format(self, column: Column) -> Optional[str]:
        """Infers datetime format (basic heuristic)."""
        if "date" in column.name.lower():
            return "%Y-%m-%d" 
        return None


if 'tables' not in st.session_state:
    st.session_state['tables'] = None
if 'db_name' not in st.session_state:
    st.session_state['db_name'] = None
if 'db_connection' not in st.session_state:
    st.session_state['db_connection'] = None
if 'fetch_rows' not in st.session_state:
    st.session_state['fetch_rows'] = None
if 'table_description' not in st.session_state:
    st.session_state['table_description'] = None
if 'column_description' not in st.session_state:
    st.session_state['column_description'] = None
if 'number_of_rows' not in st.session_state:
    st.session_state['number_of_rows'] = None
if 'sql_agent' not in st.session_state:
    st.session_state['sql_agent'] = None
if 'agent_query' not in st.session_state:
    st.session_state['agent_query'] = None
if 'hci_data' not in st.session_state:
    st.session_state['hci_data'] = None
if 'metadata' not in st.session_state:
    st.session_state['metadata'] = None
if 'metadata_dict' not in st.session_state:
    st.session_state['metadata_dict'] = None
if 'synthetic_data' not in st.session_state:
    st.session_state['synthetic_data'] = None
if 'stylised_hci_data' not in st.session_state:
    st.session_state['stylised_hci_data'] = None
if 'stylised_synthetic_data' not in st.session_state:
    st.session_state['stylised_synthetic_data'] = None
if 'specification' not in st.session_state:
    st.session_state['specification'] = None
if 'description' not in st.session_state:
    st.session_state['description'] = "N/A"
if 'report' not in st.session_state:
    st.session_state['report'] = None
if 'diagnostics' not in st.session_state:
    st.session_state['diagnostics'] = None


st.set_page_config(
    layout="wide",
    page_title="Synthetic Data Generator",
    page_icon="🧪",
)

st.write("### Generate and Evaluate Synthetic Data for Test Cases 🧪")

st.markdown(
"""
To get started:
- In the sidebar, upload a dataset (upto 5 tables) or connect to a database by providing the required credentials.  
- Optionally, add a business glossary for refined results.
- Specify your test scenario to fetch data from all relevant data in the database (upto 5 tables).
- Validate metadata.
- Generate synthetic data for upto 5 tables at once.
- We recommend using the **Smart Backfill** feature to enhance the synthetic data. This feature logically backfills null values in the synthetic data.
- You may also generate detailed reports to evaluate the synthetic data by utilising the **Generate Evaluation Report** feature.
- Download a comprehensive Excel workbook containing the original as well as synthetic data, alongwith the results from evaluation.
""")


with st.sidebar:
    st.markdown("## Provide Production Data")

    with st.expander("Upload Dataset"):
        file = st.file_uploader("Upload a file", type=['csv', 'xlsx'])
        if file is not None and file.name.endswith(".xlsx"):
            workbook = load_workbook(file)
            tables = st.multiselect("Select Tables (upto 5)", options=workbook.sheetnames)
            st.session_state['tables'] = tables
        business_glossary_file = st.file_uploader("Optionally, upload a business glossary", type=['xlsx'])
        if business_glossary_file:
            table_description, column_description = get_business_glossary(business_glossary_file)
            st.session_state['column_description'] = column_description
            st.session_state['table_description'] = table_description   

    st.markdown("## OR")

    with st.expander("Connect to Database"):
        db_type = st.selectbox("Database Type", ["PostgreSQL", "MySQL", "Microsoft SQL Server", "OracleDB", "SQLite"])

        connection_details = {}
        if db_type == "SQLite":
            connection_details['database'] = st.text_input("Database File Path")
        else:
            connection_details['host'] = st.text_input("Host")
            connection_details['username'] = st.text_input("Username")
            connection_details['password'] = st.text_input("Password", type="password")
            connection_details['database'] = st.text_input("Database Name")
            # connection_details['business_glossary'] = st.text_input("Path to Business Glossary")
            if db_type in ["PostgreSQL", "MySQL"]:
                connection_details['port'] = st.text_input("Port", value='5432' if db_type == 'PostgreSQL' else '3306')
            if db_type == "OracleDB":
                connection_details['port'] = st.text_input("Port", value='1521')
                connection_details['service'] = st.text_input("Service Name")
            if db_type == "Microsoft SQL Server":
                connection_details['authentication'] = st.selectbox("Authentication Type", ["Windows Authentication", "Others"])
            
        business_glossary_file = st.file_uploader("Optionally, upload a Business glossary", type=['xlsx'])

        if st.button("Connect to DB"):
            config = {
                "rdbms": db_type,
                **connection_details
            }
            st.session_state['db_name'] = connection_details['database']
            conn_string = get_connection_string(config)
            if conn_string is not None:
                try:
                    schemas = get_schemas(conn_string)
                    # table_description, column_description = get_business_glossary(connection_details['business_glossary'].strip('"')) if connection_details['business_glossary'].strip() != "" else (None, None)
                    if business_glossary_file:
                        table_description, column_description = get_business_glossary(business_glossary_file)
                        st.session_state['column_description'] = column_description
                        st.session_state['table_description'] = table_description   

                    db = SQLDatabase.from_uri(conn_string, schemas=schemas, view_support=False, table_description=table_description, column_description=column_description)
                    st.session_state['db_connection'] = db

                    st.success("Connected to Database!")

                except Exception as e:
                    st.error(f"Connection failed or SQL Agent initialization failed: {e}")
            else:
                st.error("Please provide valid connection details")

        if st.session_state['db_connection'] is not None:
            fetch_rows = st.selectbox("No. of rows of data to be fetched for each table", options=[100, 500, 1000, 5000, 10000, 20000, 50000, 100000])
            st.session_state['fetch_rows'] = fetch_rows

    number_of_rows = st.selectbox("No. of rows of synthetic data to be generated", options=[100, 500, 1000, 5000, 10000, 20000, 50000, 100000])
    st.session_state['number_of_rows'] = number_of_rows

if file is not None:
    hci_data = handle_upload(file, st.session_state['tables'])
    st.session_state['hci_data'] = hci_data
    st.session_state['stylised_hci_data'] = {k: v.style.highlight_null(color='yellow') for k, v in hci_data.items()}

if st.session_state['hci_data'] is not None and st.session_state['metadata'] is None:
    metadata = generate_metadata(st.session_state['hci_data'])
    st.session_state['metadata'] = metadata

if st.session_state['db_connection'] is not None:
    st.markdown("### Fetch Relevant Tables from Database")
    specification = st.text_area("Please describe the Test Scenario:")
    st.session_state['specification'] = specification.strip()
    if st.session_state['specification'] != "":
        if st.button("Fetch Data from DB"):  
            with st.spinner("Generating SQL Query and Fetching Data..."):
                try:
                    db = st.session_state['db_connection']
                    dialect = db.dialect
                    fetch_rows = st.session_state['fetch_rows']
                    agent_executor = create_sql_agent(llm, db=db, agent_type="openai-tools", verbose=True, max_iterations=10)

                    sql_agent_prompt = PromptTemplate.from_template(
                    """
                    You are a SQL expert. Based on the provided database schema and the user's test scenario specification,
                    generate a set of SQL queries to fetch the data relevant to the test scenario from relevant tables, in a
                    table-wise manner. You must infer the intent and objective of the test case to effectively query and
                    retrieve the data.
                    The query should accurately follow the syntax of the {dialect} dialect.
                    You should only query tables that are relevant to the test scenario. You should query for all columns in
                    each relevant table.
                    The queries should fetch no more than {fetch_rows} rows from each relevant table.
                    The queries may join multiple tables across multiple schemas to robustly implement complex conditions and
                    clauses but the data retrieved should be from a maximum of the top 5 most relevant tables.

                    Queries should be table-wise, for eg,
                    if you have been asked to fetch 100 rows of data from each relevant table, instead of the query being:
                    ```sql
                    SELECT TOP 100 t1.product_name, t1.amount, t2.name, t2.gender
                    FROM schema1.table1 as t1
                    JOIN schema.1table2 as t2
                    ON t1.some_id=t2.some_id
                    WHERE t2.gender = "Male";
                    ```

                    it should be:
                    ```sql
                    SELECT TOP 100 t1.product_name, t1.amount
                    FROM schema1.table1 as t1
                    JOIN schema.1table2 as t2
                    ON t1.some_id=t2.some_id
                    WHERE t2.gender = "Male";

                    SELECT TOP 100 t2.name, t2.gender
                    FROM schema1.table1 as t1
                    JOIN schema.1table2 as t2
                    ON t1.some_id=t2.some_id
                    WHERE t2.gender = "Male";
                    ```                

                    You will validate the syntax of the query using the query_sql_checker_tool.
                    You will check whether the query is retrieving data using the query_sql_database_tool.
                    In case data is retrieved, it will be indicated that the query will data alongwith the number of rows of data
                    that will be retrieved.
                    In case the indication states that no data is retrieved, you will not include that query in the final response.
                    Your output will be the SQL query to fetch the data needed for the test scenario. 
                    Do not provide any explanation or context.
                    If none of the queries retrieve no data, respond with "No data found".

                    Test Scenario:
                    {specification}           
                    """
                    )

                    agent_result = agent_executor.invoke(sql_agent_prompt.format(dialect=dialect, fetch_rows=fetch_rows, specification=specification))
                    agent_query = agent_result['output']
                    print(agent_result, "\n\n", agent_query)
                    if agent_query != "Agent stopped due to max iterations.":
                        start_index = agent_query.find("```sql")
                        end_index = agent_query.rfind("```")
                        agent_query = agent_query[start_index:end_index]
                        st.session_state['agent_query'] = agent_query.replace("```sql", "").replace("```", "")
                        print(start_index, end_index, "\n\n", st.session_state['agent_query'])
                        db_data = db.table_data(command=str(st.session_state['agent_query']))
                        if db_data is not None:
                            metadata = db.table_metadata 
                            st.session_state['hci_data'] = db_data
                            st.session_state['metadata'] = Metadata.load_from_dict(metadata)
                            st.session_state['stylised_hci_data'] = {k: v.style.highlight_null(color='yellow') for k, v in db_data.items()}

                        else:
                            st.warning("No data found")

                    else:
                        st.warning("Attempts exhausted. Please try again.")

                except Exception as e:
                    st.error(f"Error generating SQL query or fetching data: {e}")
                    pass

if st.session_state['metadata'] is not None:
    st.subheader("Validate Metadata", divider='violet')
    metadata_dict = st.session_state['metadata'].to_dict()
    st.markdown("#### Modify Metadata")
    with st.form("metadata_edit_form"):
        edited_metadata = {}
        table_metadata = st.text_area(f"Edit tables (data types, PII masking)", value=str(metadata_dict['tables']), height=400)
        relationship_metadata = st.text_area(f"Edit table relationships", value=str(metadata_dict['relationships']), height=200)
        edited_metadata['METADATA_SPEC_VERSION'] = metadata_dict['METADATA_SPEC_VERSION']
        
        submitted = st.form_submit_button("Save Changes")
        
        if submitted:
            try:
                edited_metadata['relationships'] = json.loads(relationship_metadata.replace("\'", "\""))
                edited_metadata['tables'] = json.loads(table_metadata.replace("\'", "\""))
                st.session_state['metadata_dict'] = edited_metadata
                st.session_state['metadata'] = Metadata.load_from_dict(edited_metadata)
                st.success("Metadata updated successfully!")
            except Exception as e:
                st.error(e)

    st.markdown("#### View Metadata")
    st.json(st.session_state['metadata'].to_dict(), expanded=False)

if st.session_state['hci_data'] is not None:
    st.subheader("Preview Original Data", divider='violet')
    for table_name, data in st.session_state['stylised_hci_data'].items():
        with st.expander(table_name):
            st.dataframe(data)

if st.session_state['hci_data'] is not None:
    if st.button("Generate Synthetic Data"):
        with st.spinner("Synthesizing Data..."):
            metadata = st.session_state['metadata']
            cleaned_data = drop_unknown_references(st.session_state['hci_data'], metadata)
            synthetic_data = model_hma_synthesizer(cleaned_data, metadata, st.session_state['number_of_rows'])
            st.session_state['synthetic_data'] = synthetic_data
            st.session_state['stylised_synthetic_data'] = {k: v.style.highlight_null(color='yellow') for k, v in synthetic_data.items()} # Stylise each df

        st.subheader("Synthetic Data", divider="violet")
        for table_name, data in st.session_state['stylised_synthetic_data'].items():
            with st.expander(table_name):
                st.dataframe(data)

if st.session_state['synthetic_data'] is not None:
    if st.button("Download Synthetic Data"):
        file_label = file.name.replace(".csv", "").replace(".xlsx", "") if file else st.session_state['db_name']
        if not os.path.exists(f"./output/synthetic data/{file_label}"):
            os.makedirs(f"./output/synthetic data/{file_label}")

        with pd.ExcelWriter(f"./output/synthetic data/{file_label}/{file_label} synthetic data.xlsx") as writer:
            for table_name, data in st.session_state['synthetic_data'].items():
                data.to_excel(writer, sheet_name=table_name, index=False)
        
        st.success("The synthetic data has been saved in the 'output/synthetic data' directory")

if st.session_state['synthetic_data'] is not None:
    if st.button("Smart Backfill"):
        if st.session_state['specification'] is None:
            specification = st.text_area("Please describe the Test Scenario for Smart Backfill:")
            st.session_state['specification'] = specification
        if st.session_state['table_description'] is not None and st.session_state['column_description'] is not None:
            st.session_state['description'] = {
                table: f"{st.session_state['table_description'][table]}\n\n{st.session_state['column_description'][table]}" 
                for table in st.session_state['synthetic_data'].keys() 
                if table in st.session_state['table_description'] 
            }
        if st.session_state['description'] == "N/A":
            description = st.text_area("Optionally, provide description of fields")
            if description is not None:
                st.session_state['description'] = description

        if st.session_state['specification'] is not None:
            # if st.button("Proceed"):
                with st.spinner("Backfilling Nulls..."):
                    try:
                        backfilled_synthetic_data = {}
                        total_nulls_before = 0
                        total_nulls_after = 0
                        for table_name, data in st.session_state['synthetic_data'].items():
                            nulls_before, nulls_after, backfilled_data = backfill_synthetic_data(
                                data,
                                st.session_state['specification'],
                                str(st.session_state['description'])
                            )
                            backfilled_synthetic_data[table_name] = backfilled_data
                            total_nulls_before += nulls_before
                            total_nulls_after += nulls_after

                        st.session_state['synthetic_data'] = backfilled_synthetic_data
                        st.session_state['stylised_synthetic_data'] = {k: v.style.highlight_null(color='yellow') for k, v in backfilled_synthetic_data.items()}

                        st.subheader("Synthetic Data", divider="violet")
                        col1, col2 = st.columns(2)
                        col1.metric("Total Nulls Before", total_nulls_before)
                        col2.metric("Total Nulls Now", total_nulls_after)

                        for table_name, data in st.session_state['stylised_synthetic_data'].items():
                            with st.expander(table_name):
                                st.dataframe(data)

                    except Exception as e:
                        st.error(e)

if st.session_state['synthetic_data'] is not None:
    if st.button("Evaluate Synthetic Data"):
        with st.spinner('Generating Report...'):
            try:
                file_label = file.name.replace(".csv", "").replace(".xlsx", "") if file else st.session_state['db_name']
                if not os.path.exists(f"./output/synthetic data/{file_label}"):
                    os.makedirs(f"./output/synthetic data/{file_label}")

                for table_name, data in st.session_state['synthetic_data'].items():
                    table_evaluator_instance = TableEvaluator(
                        st.session_state['hci_data'][table_name], 
                        st.session_state['synthetic_data'][table_name], 
                        # cat_cols=st.session_state['hci_data'][table_name].select_dtypes(include=['category']).columns.tolist()
                    )
                    table_evaluator_instance.visual_evaluation(save_dir=f"./output/synthetic data/{file_label}/{table_name}")
            except Exception as e:
                # st.warning(f"TableEvaluator - {e}")
                pass

            try:
                report = QualityReport()
                diagnostics = DiagnosticReport()
                report.generate(st.session_state['hci_data'], st.session_state['synthetic_data'], st.session_state['metadata'].to_dict())
                diagnostics.generate(st.session_state['hci_data'], st.session_state['synthetic_data'], st.session_state['metadata'].to_dict())
                report_properties = report.get_properties()
                diagnostics_properties = diagnostics.get_properties()

                empty_row = pd.DataFrame(columns=report.get_details(property_name="Column Shapes").columns, index=[0])
                st.session_state['report'] = pd.concat([report.get_details(property_name="Column Shapes"), empty_row, report.get_details(property_name="Column Pair Trends")], ignore_index=True)
                empty_row = pd.DataFrame(columns=diagnostics.get_details(property_name="Data Validity").columns, index=[0])
                st.session_state['diagnostics'] = pd.concat([diagnostics.get_details(property_name="Data Validity"), empty_row, diagnostics.get_details(property_name="Data Structure")], ignore_index=True)

            except Exception as e:
                st.warning(f"SDMetrics - {e}")

            st.subheader("Synthetic Data", divider="violet")
            for table_name, data in st.session_state['stylised_synthetic_data'].items():
                with st.expander(table_name):
                    st.dataframe(data)
                
            try:
                st.subheader("Evaluation Result", divider="violet")
                st.write("The report has been structured into qualitative analysis followed by quantitative indicators of the original and synthetic datasets")
                st.text("\n\n")
                st.write("### Qualitative Analysis")
                st.text("\n\n")
                st.write("**Quality Metrics**: The following metrics evaluate how well the synthetic data captures the mathematical and categorical properties in the original dataset. Visualizations complementing this analysis have been provided further down the report. A more detailed analysis can be downloaded.")
                col1, col2, col3 = st.columns(3)
                col1.metric("Overall Score", f"{"{:.2f}".format(report.get_score() * 100)}%")
                col2.metric("Column Shapes Score", f"{"{:.2f}".format(report_properties.loc[0, 'Score'] * 100)}%")
                col3.metric("Column Pair Trends Score", f"{"{:.2f}".format(report_properties.loc[1, 'Score'] * 100)}%")
                st.text("\n\n")
                st.write("**Diagnostic Metrics**: The following metrics will provide a general sense of the strengths and weaknesses of the synthetic data on the basis of relationship validity and structure. These scores are computed on the basis of boundary and category adherance, key uniqueness, cardinality adherence, overall table structure, missing value similarity, statistical similarity, coverage of range and categories as well as sequence length similarity. A more detailed analysis can be downloaded.")
                col1, col2, col3 = st.columns(3)
                col1.metric("Overall Score", f"{"{:.2f}".format(diagnostics.get_score() * 100)}%")
                col2.metric("Data Validity Score", f"{"{:.2f}".format(diagnostics_properties.loc[0, 'Score'] * 100)}%")
                col3.metric("Data Structure Score", f"{"{:.2f}".format(diagnostics_properties.loc[1, 'Score'] * 100)}%")
                st.text("\n\n")
                for table_name in st.session_state['synthetic_data'].keys():
                    if os.listdir(f"./output/synthetic data/{file_label}/{table_name}"):
                        st.write(f"### Quantitative Analysis for Synthetic Data of {table_name}")
                        with st.expander(table_name):
                            if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/mean_std.png"):
                                st.write("The following visualises a high-level overview of trends in the numerical data from the original dataset as well as synthetic dataset")
                                st.image(f"./output/synthetic data/{file_label}/{table_name}/mean_std.png")
                            st.text("\n\n\n")
                            if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/cumsums.png"):
                                st.write("A comparison between the cumulative sums of each feature in numerical data from the real and synthetic dataset summarizes the extent to which the synthetic data abstracts trends in the original dataset")
                                st.image(f"./output/synthetic data/{file_label}/{table_name}/cumsums.png")
                            st.text("\n\n\n")
                            if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/distributions.png"):
                                st.write("A comparison between distribution of data points per feature summarizes how closely the synthetic data mimics the data distribution of the original dataset")
                                st.image(f"./output/synthetic data/{file_label}/{table_name}/distributions.png")
                            st.text('\n')
                            if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/correlation_difference.png"):
                                st.write("The following set of heatmaps summarise how well the synthetic data captures the correlations between the features in the original dataset")
                                st.image(f"./output/synthetic data/{file_label}/{table_name}/correlation_difference.png")
                            st.text("\n\n\n")
                            if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/pca.png"):
                                st.write("Further analysis of the differences in the feature-wise correlations has been represented as the first two principal components of the original and synthetic data")
                                st.image(f"./output/synthetic data/{file_label}/{table_name}/pca.png")
                st.write("-------")

            except Exception as e:
                st.error(f"An error occurred: Displaying - {e}")

if st.session_state['synthetic_data'] is not None and st.session_state['report'] is not None and st.session_state['diagnostics'] is not None:
    if st.button("Download with Evaluation Report"):
        with st.spinner("Compiling..."):
            try:
                file_label = file.name.replace(".csv", "").replace(".xlsx", "") if file else st.session_state['db_name']
                with pd.ExcelWriter(f"./output/synthetic data/{file_label}/{file_label} synthetic data.xlsx", engine='openpyxl') as writer:
                    for table_name in st.session_state['synthetic_data'].keys():
                        st.session_state['hci_data'][table_name].to_excel(writer, sheet_name=f"Original Data in {table_name}", index=False)
                        st.session_state['synthetic_data'][table_name].to_excel(writer, sheet_name=f"Synthetic Data from {table_name}", index=False)
                    
                    st.session_state['report'].to_excel(writer, sheet_name="Quality Metrics", index=False)
                    st.session_state['diagnostics'].to_excel(writer, sheet_name="Diagnostic Metrics", index=False)

                wb = load_workbook(f"./output/synthetic data/{file_label}/{file_label} synthetic data.xlsx")
                for table_name in st.session_state['synthetic_data'].keys():
                    if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/mean_std.png"):
                        ws = wb.create_sheet(f"Mean and Standard Deviation for {table_name}")
                        image = Image(f"./output/synthetic data/{file_label}/{table_name}/mean_std.png")
                        ws.add_image(image, 'A1')
                    if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/cumsums.png"):
                        ws = wb.create_sheet(f"Cumulative Sums per Feature for {table_name}")
                        image = Image(f"./output/synthetic data/{file_label}/{table_name}/cumsums.png")
                        ws.add_image(image, 'A1')
                    if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/distributions.png"):
                        ws = wb.create_sheet(f"Distribution per Feature for {table_name}")
                        image = Image(f"./output/synthetic data/{file_label}/{table_name}/distributions.png")
                        ws.add_image(image, 'A1')
                    if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/correlation_difference.png"):
                        ws = wb.create_sheet(f"Correlation and Difference for {table_name}")
                        image = Image(f"./output/synthetic data/{file_label}/{table_name}/correlation_difference.png")
                        ws.add_image(image, 'A1')
                    if os.path.exists(f"./output/synthetic data/{file_label}/{table_name}/pca.png"):
                        ws = wb.create_sheet(f"Principal Component Analysis for {table_name}")
                        image = Image(f"./output/synthetic data/{file_label}/{table_name}/pca.png")
                        ws.add_image(image, 'A1')

                wb.save(f"./output/synthetic data/{file_label}/{file_label} synthetic data.xlsx")

                st.success("The synthetic data alongwith a detailed evaluation report has been saved in the 'output/synthetic data' directory")
            except Exception as e:
                st.error(f"An error occured: {e}")